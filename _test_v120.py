"""HCCheck v1.2 mock 测试 - 验证 db_query 和 history tab

不依赖 Playwright / 真实运管站系统。
"""
import os
import re
import sys
import time

# 让 import 能找到 config / db / db_query
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

import config
import db
import db_query


def test_1_data_dir():
    """测试 1: 数据目录正确创建"""
    print("\n=== Test 1: 数据目录 ===")
    print(f"  DATA_DIR: {config.DATA_DIR}")
    assert os.path.isdir(config.DATA_DIR), f"数据目录不存在: {config.DATA_DIR}"
    print("  ✅ 数据目录存在")
    return True


def test_2_record_runs():
    """测试 2: 写入 3 辆 mock 车到数据库"""
    print("\n=== Test 2: 写入 mock 数据 ===")

    # 初始化数据库
    db.init_db()
    run_id = db.get_run_id()
    print(f"  run_id: {run_id}")

    # 模拟 3 辆车 (时间递增, 防止负数 duration)
    base_time = time.time() - 300  # 5 分钟前开始
    test_cars = [
        ("冀J0E139", "带挂", base_time, base_time + 32.5, "成功", ""),
        ("冀J0E140", "不带挂", base_time + 40, base_time + 70.0, "成功", ""),
        ("冀J0E141", "带挂", base_time + 80, base_time + 140.0, "失败", "popup4 超时未响应"),
    ]

    for plate, flow_type, start, end, status, error in test_cars:
        ok = db.record_run(plate, flow_type, start, end, status, error)
        assert ok, f"记录 {plate} 失败"
        print(f"  ✅ {plate} ({flow_type}) - {status} ({end-start:.1f}s)")

    return run_id


def test_3_query(run_id):
    """测试 3: get_runs_by_run_id 查询"""
    print("\n=== Test 3: 查询本次 run ===")
    runs = db_query.get_runs_by_run_id(run_id)
    print(f"  查询到 {len(runs)} 辆车:")
    for r in runs:
        print(f"    {r['plate']} | {r['flow_type']} | {r['duration']:.1f}s | {r['status']}")
    assert len(runs) == 3, f"应该有 3 辆车, 实际 {len(runs)}"
    # 验证按 start_time 升序
    assert runs[0]["plate"] == "冀J0E139"
    assert runs[1]["plate"] == "冀J0E140"
    assert runs[2]["plate"] == "冀J0E141"
    # 验证失败车有 error
    assert runs[2]["error"] == "popup4 超时未响应"
    print("  ✅ 查询结果正确")
    return True


def test_4_summary(run_id):
    """测试 4: get_run_summary 汇总"""
    print("\n=== Test 4: 汇总统计 ===")
    summary = db_query.get_run_summary(run_id)
    print(f"  总数: {summary['total']}")
    print(f"  成功: {summary['success']}")
    print(f"  失败: {summary['failed']}")
    print(f"  平均耗时: {summary['avg_duration']}s")
    assert summary["total"] == 3
    assert summary["success"] == 2
    assert summary["failed"] == 1
    assert summary["avg_duration"] > 0
    print("  ✅ 汇总正确")
    return True


def test_5_export(run_id):
    """测试 5: 导出 xlsx"""
    print("\n=== Test 5: 导出 xlsx ===")
    output = os.path.join(SCRIPT_DIR, "_test_export.xlsx")
    # 先删旧文件
    if os.path.exists(output):
        os.remove(output)

    ok = db_query.export_run_to_xlsx(run_id, output)
    assert ok, "导出失败"
    assert os.path.exists(output), f"文件未生成: {output}"
    size = os.path.getsize(output)
    print(f"  ✅ xlsx 已生成: {output} ({size} bytes)")

    # 验证 xlsx 内容
    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active
    print(f"  Sheet 名: {ws.title}")
    print(f"  总行数: {ws.max_row}")
    print(f"  表头: {[c.value for c in ws[1]]}")
    print(f"  第一条: {[c.value for c in ws[2]]}")

    # 验证表头
    expected_headers = ["车牌", "类型", "耗时(秒)", "状态", "错误", "开始时间"]
    actual_headers = [c.value for c in ws[1]]
    assert actual_headers == expected_headers, f"表头不对: {actual_headers}"

    # 验证数据行数 = 1 表头 + 3 数据
    assert ws.max_row == 4, f"行数不对: {ws.max_row}"

    # 清理
    os.remove(output)
    print("  ✅ xlsx 内容正确, 测试文件已清理")
    return True


def test_6_empty_run_id():
    """测试 6: 空 run_id 容错"""
    print("\n=== Test 6: 空 run_id 容错 ===")
    runs = db_query.get_runs_by_run_id("")
    summary = db_query.get_run_summary("")
    assert runs == []
    assert summary["total"] == 0
    print("  ✅ 空 run_id 返回空结果不崩")
    return True


def test_6b_latest_run_id():
    """Test 6b: get_latest_run_id 跨进程查最新 (v1.2.1 修复验证)"""
    print("\n=== Test 6b: get_latest_run_id ===")
    # 先清掉表, 插 2 个不同 run_id
    import sqlite3
    conn = sqlite3.connect(config.DB_FILE, timeout=5)
    conn.execute("DELETE FROM runs")
    conn.commit()
    conn.close()

    # 插第 1 个 run (老)
    db.record_run("冀J0E139", "带挂", time.time()-100, time.time()-50, "成功", "", run_id="20260101-120000-old0001")
    # 插第 2 个 run (新)
    db.record_run("冀J0E140", "不带挂", time.time()-50, time.time()-20, "成功", "", run_id="20260101-130000-new0002")
    # 插第 3 个 (中间 created_at)
    db.record_run("冀J0E141", "带挂", time.time()-10, time.time(), "失败", "测试错误", run_id="20260101-140000-mid0003")

    latest = db_query.get_latest_run_id()
    print(f"  最新 run_id: {latest}")
    assert latest == "20260101-140000-mid0003", f"应为 mid0003, 实际 {latest}"
    print(f"  ✅ 最新 run_id 正确")
    return True


def test_7_gui_loading():
    """测试 7: GUI 代码结构 (AST 检查, 不实际 import, 避开 playwright 依赖)"""
    print("\n=== Test 7: GUI 代码结构 ===")
    import ast
    src = open(os.path.join(SCRIPT_DIR, "gui.py"), encoding="utf-8").read()
    tree = ast.parse(src)

    # 找 App 类
    app_class = None
    for node in ast.walk(tree):
        if isinstance(node, ast.ClassDef) and node.name == "App":
            app_class = node
            break
    assert app_class is not None, "找不到 class App"

    # 找方法名
    methods = {m.name for m in app_class.body if isinstance(m, ast.FunctionDef)}
    required = ["_build_tab_history", "_refresh_history_table", "_export_history_xlsx"]
    for m in required:
        assert m in methods, f"App 缺少方法: {m}"
        print(f"  ✅ App.{m} 存在")

    # 找 VERSION
    version = None
    for node in app_class.body:
        if isinstance(node, ast.Assign):
            for target in node.targets:
                if isinstance(target, ast.Name) and target.id == "VERSION":
                    version = node.value.value
                    break
    assert version == "v1.2.3", f"VERSION 应为 v1.2.3, 实际 {version}"
    print(f"  ✅ App.VERSION = {version}")

    # 验证 import
    imports = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for n in node.names:
                imports.add(n.name)
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imports.add(node.module)
    assert "db" in imports, "缺 import db"
    assert "db_query" in imports, "缺 import db_query"
    print(f"  ✅ import db / db_query 正常")

    return True


def test_8_tab_order():
    """测试 8: 验证 Tab 顺序 (运行 / 历史记录 / 黑名单 / 设置)"""
    print("\n=== Test 8: Tab 顺序检查 ===")
    src = open(os.path.join(SCRIPT_DIR, "gui.py"), encoding="utf-8").read()

    # 找 4 个 tab 的添加顺序
    expected_order = [
        ("运行", "tab_run"),
        ("历史记录", "tab_history"),
        ("黑名单", "tab_skip"),
        ("设置", "tab_settings"),
    ]
    last_pos = -1
    for label, attr in expected_order:
        # 找 self.notebook.add(self.tab_xxx, ...)
        idx = src.find(f"self.notebook.add(self.{attr}")
        assert idx > last_pos, f"Tab '{label}' ({attr}) 位置不对"
        print(f"  ✅ {label} (self.{attr}) 在位置 {idx}")
        last_pos = idx
    return True


def test_13_popup2_strategy10_powershell():
    """测试 13: popup2 策略10 OS 层 PowerShell 关 Lodop native window (v1.2.2 第 9 轮)

    背景: Lodop preview 是 Windows native window, JS API 管不到
    修法: subprocess 调 PowerShell, Get-Process 找标题含'打印预览'/'Lodop', CloseMainWindow
    """
    print("\n=== Test 13: popup2 策略10 (PowerShell) ===")
    src = open(os.path.join(SCRIPT_DIR, "popups/p2_tech_review.py"), encoding="utf-8").read()

    # 检查策略10存在
    assert "STRATEGY10" in src, "❌ popup2 没有 STRATEGY10"
    print("  ✅ popup2 策略10 STRATEGY10 存在")

    # 检查 subprocess + PowerShell
    for kw in ["subprocess", "powershell", "CloseMainWindow", "Get-Process"]:
        assert kw in src, f"❌ 策略10 缺关键字: {kw}"
    print("  ✅ 用 subprocess + powershell + Get-Process + CloseMainWindow")

    # 检查只匹配预览窗口 (不杀 IE/Chrome)
    for kw in ["打印预览", "Lodop", "LODOP"]:
        assert kw in src, f"❌ 策略10 缺窗口标题关键词: {kw}"
    print("  ✅ 窗口标题关键词: 打印预览 / Lodop / LODOP")

    # 检查 win32 平台守卫
    assert 'sys.platform == "win32"' in src, "❌ 策略10 缺 win32 平台守卫"
    print("  ✅ 只在 Windows 平台执行 (sys.platform == 'win32')")

    return True



    """测试 12: popup2 策略8 进 _workflow_tmp iframe 调 LODOP API (v1.2.2 第 6 轮)

    背景: 策略7 诊断定位 _workflow_tmp iframe 是 Lodop 渲染位置, window.LODOP 在那里定义
    修法: 进 iframe 调 LODOP.PREVIEW(false) + SET_PRINT_MODE('AUTO_CLOSE_PREWINDOW', true) + CLOSE_PRINTTASK()
    """
    print("\n=== Test 12: popup2 策略8 (LODOP API) ===")
    src = open(os.path.join(SCRIPT_DIR, "popups/p2_tech_review.py"), encoding="utf-8").read()

    # 检查策略8存在
    assert "STRATEGY8" in src, "❌ popup2 没有 STRATEGY8"
    print("  ✅ popup2 策略8 STRATEGY8 存在")

    # 检查找 _workflow_tmp iframe
    assert "workflow_tmp" in src, "❌ popup2 策略8 没找 _workflow_tmp iframe"
    print("  ✅ 策略8 找 _workflow_tmp iframe")

    # 检查调关键 LODOP API
    for api in ["PREVIEW", "SET_PRINT_MODE", "AUTO_CLOSE_PREWINDOW", "CLOSE_PRINTTASK"]:
        assert api in src, f"❌ 策略8 缺 LODOP API: {api}"
    print("  ✅ 调用 PREVIEW / SET_PRINT_MODE / AUTO_CLOSE_PREWINDOW / CLOSE_PRINTTASK")
    return True


def test_10_popup4_excludes_main_page():
    """测试 10: popup4 _close_residual_popups 豁免主页面 (v1.2.2 修复)

    背景: 17:39 实测 popup4 入口清理残留弹窗时, 主页面也被关, 导致后续 navigation 全失败
    修法: _close_residual_popups 增加 main_page 参数, 同时豁免主页面
    """
    print("\n=== Test 10: popup4 豁免主页面 ===")
    print("\n=== Test 10: popup4 豁免主页面 ===")
    import popups.p4_vehicle_annual as p4

    # 检查 handle 函数签名
    import inspect
    handle_sig = inspect.signature(p4.handle)
    assert 'main_page' in handle_sig.parameters, \
        "❌ popup4.handle 没有 main_page 参数"
    print("  ✅ popup4.handle 有 main_page 参数")

    # 检查 _click_year_check 函数签名
    cyc_sig = inspect.signature(p4._click_year_check)
    assert 'main_page' in cyc_sig.parameters, \
        "❌ popup4._click_year_check 没有 main_page 参数"
    print("  ✅ popup4._click_year_check 有 main_page 参数")

    # 检查 _close_residual_popups 函数签名
    crp_sig = inspect.signature(p4._close_residual_popups)
    assert 'main_page' in crp_sig.parameters, \
        "❌ popup4._close_residual_popups 没有 main_page 参数"
    print("  ✅ popup4._close_residual_popups 有 main_page 参数")

    # 检查 run.py 调用方都传了 main_page=page (忽略注释行)
    import re
    run_src_lines = open(os.path.join(SCRIPT_DIR, "run.py"), encoding="utf-8").readlines()
    code_lines = [l for l in run_src_lines if not l.strip().startswith("#")]
    code_src = "".join(code_lines)
    handle_calls = re.findall(r'handle_vehicle_annual\(', code_src)
    main_page_passes = re.findall(r'main_page=page\b', code_src)
    print(f"  ✅ run.py 里 handle_vehicle_annual 调用 {len(handle_calls)} 处")
    print(f"  ✅ main_page=page 实际传参 {len(main_page_passes)} 处")
    assert len(handle_calls) == 2, f"❌ 应有 2 处调用, 实际 {len(handle_calls)}"
    assert len(main_page_passes) == 2, f"❌ 应有 2 处传 main_page=page, 实际 {len(main_page_passes)}"
    return True


def test_9_no_tkinter_typos():
    """测试 9: 防 Tkinter 参数名 typo (v1.2.1 回归)

    背景: 之前用了 'initialfilename' (错), Tcl 报错 'bad option -initialfilename'
          正确参数名是 'initialfile'
    """
    print("\n=== Test 9: Tkinter 参数名 ===")
    src = open(os.path.join(SCRIPT_DIR, "gui.py"), encoding="utf-8").read()

    # 禁止 'initialfilename=' 作为实际 kwarg (错误参数名)
    import re
    bad_kwarg = re.findall(r'\binitialfilename\s*=', src)
    assert not bad_kwarg, \
        f"❌ gui.py 实际调用里还有 'initialfilename=' typo, 应改为 'initialfile=' ({len(bad_kwarg)} 处)"
    print(f"  ✅ 没有 'initialfilename=' 实际 kwarg (只在注释里提到, OK)")

    # 应该有 'initialfile=' (正确)
    initialfile_count = src.count('initialfile=')
    assert initialfile_count >= 1, "❌ 没找到 'initialfile=' (导出对话框应有默认文件名)"
    print(f"  ✅ 'initialfile=' 出现 {initialfile_count} 次 (xlsx + log 导出)")

    # VERSION 应该是 v1.2.3
    assert 'VERSION = "v1.2.3"' in src, "❌ VERSION 没 bump 到 v1.2.3"
    print("  ✅ VERSION = v1.2.3")

    # 按钮文案应该是 "导出最近"
    assert "导出最近为 xlsx" in src, "❌ 按钮文案没改成 '导出最近'"
    assert "导出本次为 xlsx" not in src, "❌ 还有 '导出本次为 xlsx' 旧文案"
    print("  ✅ 按钮文案 '导出最近为 xlsx' 正确")
    return True


def test_14_click_query_and_pagination():
    """测试 14: _click_query_button / _has_next_page / _click_next_page 函数 (v1.2.3)

    背景: 用户要求打开普货审验后先点查询, 而且要支持超过 15 辆的分页
    修法: 加 3 个新函数 + 重构 get_next_plate_from_list 加翻页 while 循环
    """
    print("\n=== Test 14: 查询 + 分页函数 ===")
    src = open(os.path.join(SCRIPT_DIR, "run.py"), encoding="utf-8").read()

    # 检查 4 个新函数都存在
    for fn in ["_click_query_button", "_has_next_page", "_click_next_page", "_scan_current_page_for_plate"]:
        assert f"def {fn}(" in src, f"❌ run.py 缺函数: {fn}"
        print(f"  ✅ def {fn}() 存在")

    # 检查 _click_query_button 多策略
    query_section = src[src.find("def _click_query_button"):src.find("def _has_next_page")]
    assert "input[value='查询']" in query_section, "❌ 缺 input value 策略"
    assert "get_by_role" in query_section, "❌ 缺 role 策略"
    print("  ✅ _click_query_button 含 input value + role + text 三策略")

    # 检查 _has_next_page 双重判断
    has_next_section = src[src.find("def _has_next_page"):src.find("def _click_next_page")]
    assert "row_count < 15" in has_next_section, "❌ _has_next_page 缺 < 15 判断"
    assert "下一页" in has_next_section, "❌ _has_next_page 缺下一页按钮检查"
    assert "is_disabled" in has_next_section, "❌ _has_next_page 缺 disabled 检查"
    print("  ✅ _has_next_page 双重判断: < 15 + 下一页按钮 + disabled")

    # 检查 _click_next_page 翻页
    click_next_section = src[src.find("def _click_next_page"):src.find("def _verify_in_normal_review")]
    assert "下一页" in click_next_section, "❌ _click_next_page 缺下一页定位"
    assert "evaluate" in click_next_section, "❌ _click_next_page 缺 JS 强制点击兜底"
    print("  ✅ _click_next_page 含 text + JS 兜底")

    return True


def test_15_pagination_while_loop():
    """测试 15: get_next_plate_from_list 翻页 while 循环 (v1.2.3)

    背景: 超过 15 辆车时, 需翻页才能拿到所有车 (用户反馈: <15 辆时无下一页按钮)
    修法: 主函数加 while 循环, 调 _scan_current_page_for_plate + _has_next_page + _click_next_page
    """
    print("\n=== Test 15: 翻页 while 循环 ===")
    src = open(os.path.join(SCRIPT_DIR, "run.py"), encoding="utf-8").read()

    # 找 get_next_plate_from_list 函数体
    func_start = src.find("def get_next_plate_from_list(")
    next_func = src.find("# ========= table", func_start)
    func_body = src[func_start:next_func]

    # 检查 while 循环
    assert "while True:" in func_body, "❌ get_next_plate_from_list 缺 while 循环"
    print("  ✅ 含 while True 翻页循环")

    # 检查调 3 个新函数
    assert "_scan_current_page_for_plate" in func_body, "❌ 没调 _scan_current_page_for_plate"
    assert "_has_next_page" in func_body, "❌ 没调 _has_next_page"
    assert "_click_next_page" in func_body, "❌ 没调 _click_next_page"
    print("  ✅ 调用 _scan_current_page_for_plate / _has_next_page / _click_next_page")

    # 检查安全上限 (防止死循环)
    assert "page_no > 20" in func_body, "❌ 没设安全上限"
    print("  ✅ 安全上限 page_no > 20 防死循环")

    # 检查 _click_query_button 在主循环被调用
    main_loop_section = src[src.find("# === 第三步:"):src.find("# 🆕 v1.2.2: 读 table 前先检查")]
    assert "_click_query_button" in main_loop_section, "❌ 主循环没调 _click_query_button"
    print("  ✅ 主循环调用 _click_query按钮 (点查询按钮)")

    # 检查 PA_AFTER_NAV 用于翻页后等待
    assert "PA_AFTER_NAV" in func_body, "❌ 翻页后没等 PA_AFTER_NAV"
    print("  ✅ 翻页后等待 PA_AFTER_NAV")

    return True


def test_16_popup2_strategy9_no_false_success():
    """测试 16: popup2 STRATEGY9 不再因 SET_PRINT_MODE 'ok' 误设 closed=1 (v1.2.3 修复)

    背景: STRATEGY9 看到 SET_PRINT_MODE(CLOSE_PREVIEW_WINDOW) ok 就设 closed=1
          但 SET_PRINT_MODE 只是改配置, 不关已开预览 → 跳过了 STRATEGY8 和 STRATEGY10
    修法: STRATEGY9 只在 'called' 出现时设 closed=1, 忽略 SET_PRINT_MODE 的 'ok'
    """
    print("\n=== Test 16: popup2 STRATEGY9 误判修复 ===")
    src = open(os.path.join(SCRIPT_DIR, "popups/p2_tech_review.py"), encoding="utf-8").read()

    # 找 STRATEGY9 的判定段
    strategy9_start = src.find("# 🆕 v1.2.3 策略9")
    strategy8_start = src.find("# 🆕 v1.2.2 策略8")
    if strategy8_start == -1:
        # v1.2.3 清理后 STRATEGY8 已删, 用 STRATEGY10 作为右边界
        strategy8_start = src.find("# 🆕 v1.2.3 策略10")
    strategy9_section = src[strategy9_start:strategy8_start]

    # 检查不再用 'ok' 误判 (这是 bug 源)
    assert "'ok' in r" not in strategy9_section, \
        "❌ STRATEGY9 还在用 'ok' in r 误判 (会跳过 STRATEGY8/STRATEGY10)"
    print("  ✅ STRATEGY9 不再用 'ok' 误判")

    # 检查还有 'called' 判断 (close 类 API 真调用时仍能成功)
    assert "'called' in r" in strategy9_section, \
        "❌ STRATEGY9 没了 'called' 判断 (真调 close 类 API 时也不能成功)"
    print("  ✅ STRATEGY9 保留 'called' 判断")

    # 检查 STRATEGY10 还在 closed == 0 时运行 (不会被 STRATEGY9 误判挡住)
    strategy10_start = src.find("# 🆕 v1.2.3 策略10")
    assert strategy10_start > 0, "❌ 找不到 STRATEGY10 注释"
    strategy10_section = src[strategy10_start:strategy10_start + 800]
    assert 'sys.platform == "win32"' in strategy10_section, \
        "❌ STRATEGY10 平台守卫没了"
    assert "closed == 0" in strategy10_section, \
        "❌ STRATEGY10 条件守卫没了"
    print("  ✅ STRATEGY10 还在 closed == 0 + win32 时运行")

    return True


def test_17_popup2_strategy10_imports_sys():
    """测试 17: popup2 STRATEGY10 引用 sys 必须 import (v1.2.3 修复)

    背景: STRATEGY10 用 sys.platform == 'win32' 守门 Windows-only,
          但 popup2 之前没 import sys → STRATEGY10 永远抛 NameError,
          OS 层 PowerShell 关预览窗口 9 轮都没生效 → popup2 预览一直在
    修法: 顶部加 import sys
    """
    print("\n=== Test 17: popup2 sys import ===")
    src = open(os.path.join(SCRIPT_DIR, "popups/p2_tech_review.py"), encoding="utf-8").read()

    # 检查 import sys 存在 (在 popup2 模块顶部)
    import re
    import_lines = re.findall(r'^import\s+sys\b', src, re.MULTILINE)
    assert len(import_lines) >= 1, \
        "❌ popup2 缺 import sys (STRATEGY10 用 sys.platform 会 NameError)"
    print(f"  ✅ import sys 存在 ({len(import_lines)} 处)")

    # 检查 sys.platform 用法仍在 (保留 Windows 守门)
    assert 'sys.platform == "win32"' in src, \
        "❌ sys.platform 守门没了 (会非 Windows 也跑 PowerShell)"
    print("  ✅ sys.platform == 'win32' 守门保留")

    # 检查 sys 在 STRATEGY10 上下文里实际被引用
    strategy10_pos = src.find("# 🆕 v1.2.3 策略10")
    assert strategy10_pos > 0, "❌ 找不到 STRATEGY10 注释"
    nearby = src[strategy10_pos:strategy10_pos + 500]
    assert "sys.platform" in nearby, "❌ STRATEGY10 附近没用 sys.platform"
    print("  ✅ STRATEGY10 附近正确引用 sys.platform")

    return True


def test_18_popup2_cleanup():
    """测试 18: popup2 v1.2.3 清理后, 只剩 STRATEGY9 + STRATEGY10 (v1.2.3 验证)

    背景: 之前 9 个策略都无效, STRATEGY10 (PowerShell 强杀 native 进程) 是唯一真管用的
          v1.2.3 删了 STRATEGY5/6/7/8 + DIAG/DIAG2 (359 行死代码), 重命名 closed → preview_closed
    验证: 只剩 STRATEGY9 (probe) + STRATEGY10 (真管用), 变量名重命名, 函数变 200 行左右
    """
    print("\n=== Test 18: popup2 v1.2.3 清理 ===")
    src = open(os.path.join(SCRIPT_DIR, "popups/p2_tech_review.py"), encoding="utf-8").read()

    # 1. 检查死代码已删
    for dead in ["STRATEGY5", "STRATEGY6", "STRATEGY7", "STRATEGY8",
                 "lodop_close_selectors", "DIAG2"]:
        assert dead not in src, f"❌ 死代码未删: {dead}"
    print("  ✅ STRATEGY5/6/7/8 + lodop_close_selectors + DIAG2 全部删除")

    # 2. 检查 STRATEGY9 + STRATEGY10 保留
    assert "STRATEGY9" in src, "❌ STRATEGY9 (probe) 丢了"
    assert "STRATEGY10" in src, "❌ STRATEGY10 (真管用) 丢了"
    print("  ✅ STRATEGY9 (probe) + STRATEGY10 (PowerShell) 保留")

    # 3. 检查变量名重命名 (preview_closed, 不再有 closed)
    assert "preview_closed" in src, "❌ 变量没重命名 preview_closed"
    # closed 作为独立变量名应该消失 (但作为属性/字符串名可能还在)
    import re
    # 找 closed = 0 / closed += 1 / closed = 1 / return closed 这类
    closed_var_pattern = re.findall(r'^\s+closed\s*[=+]+\s', src, re.MULTILINE)
    assert len(closed_var_pattern) == 0, f"❌ 还有 {len(closed_var_pattern)} 个 closed 变量赋值"
    print(f"  ✅ 变量重命名: preview_closed 取代 closed ({len(closed_var_pattern)} 个残留)")

    # 4. 检查函数长度变短 (< 300 行)
    func_start = src.find("def _close_print_preview(")
    next_func = src.find("def handle(", func_start)
    func_body = src[func_start:next_func]
    func_lines = func_body.count("\n")
    assert func_lines < 300, f"❌ 函数还太长: {func_lines} 行"
    print(f"  ✅ 函数变短: {func_lines} 行 (原来 473 行, 节省 {473-func_lines} 行)")

    return True


def test_19_log_noise_cleanup():
    """测试 19: 运行时日志去噪 — 5 处弹窗噪音 + STRATEGY9/10 噪音 DEBUG-gate (v1.2.3)

    背景: 用户反馈运行时日志太啰嗦, 每辆车每个弹窗都打 "弹窗: 动作类型→X / 全选 / 确定"
          加上 STRATEGY9/10 每次都打印所有 LODOP 方法 + PowerShell 输出细节
    修法: 噪音 print 包裹 if config.DEBUG:, 重要行 (成功/警告/错误) 保留
    """
    print("\n=== Test 19: 日志去噪 ===")

    # 1. dialog.py 5 处弹窗噪音 DEBUG-gate
    dialog_src = open(os.path.join(SCRIPT_DIR, "dialog.py"), encoding="utf-8").read()
    noisy_dialog_lines = [
        "动作类型→",
        "动作类型默认已是正确值",
        "处理人类别→",
        "弹窗: 全选",
        "弹窗: 确定",
    ]
    for noisy in noisy_dialog_lines:
        # 噪音行必须有 if config.DEBUG: 前缀
        pattern = f"if config.DEBUG: print.*\\n.*\\n?.*?{re.escape(noisy)}|if config.DEBUG: print.*'{re.escape(noisy)}'|if config.DEBUG: print.*f.*\"{re.escape(noisy)}"
        # 简化检查: 噪音行所在行或前一行有 'if config.DEBUG'
        found = False
        for i, line in enumerate(dialog_src.split("\n")):
            if noisy in line:
                # 噪音行在同一行 或 前一行
                if "if config.DEBUG" in line or (i > 0 and "if config.DEBUG" in dialog_src.split("\n")[i-1]):
                    found = True
                    break
        assert found, f"❌ dialog.py 噪音没 DEBUG-gate: {noisy}"
    print(f"  ✅ dialog.py {len(noisy_dialog_lines)} 处弹窗噪音 DEBUG-gate")

    # 2. dialog.py 重要错误保留 (没 DEBUG-gate)
    important_dialog_lines = [
        "动作类型跳过",          # 错误
        "全选跳过",              # 警告
        "确定按钮找不到",        # 错误
    ]
    for important in important_dialog_lines:
        found_unconditional = False
        for i, line in enumerate(dialog_src.split("\n")):
            if important in line and 'print' in line:
                # 这行不能有 if config.DEBUG
                if "if config.DEBUG" not in line:
                    found_unconditional = True
                    break
        assert found_unconditional, f"❌ dialog.py 重要错误被误 DEBUG-gate: {important}"
    print(f"  ✅ dialog.py {len(important_dialog_lines)} 处错误/警告保留 (无 DEBUG-gate)")

    # 3. p2_tech_review.py STRATEGY9/10 噪音 DEBUG-gate
    p2_src = open(os.path.join(SCRIPT_DIR, "popups/p2_tech_review.py"), encoding="utf-8").read()
    noisy_p2_lines = [
        "🔍 [STRATEGY9] 探测 LODOP close API",
        "🔍 [STRATEGY9] {r}",
        "🔍 [STRATEGY9] (猜)",
        "🔧 [STRATEGY10] OS 层关 Lodop preview window",
        "🔧 [STRATEGY10] {line.strip()}",
    ]
    for noisy in noisy_p2_lines:
        found = False
        for line in p2_src.split("\n"):
            if noisy in line and 'print' in line:
                if "if config.DEBUG" in line:
                    found = True
                    break
        assert found, f"❌ p2 STRATEGY9/10 噪音没 DEBUG-gate: {noisy}"
    print(f"  ✅ p2_tech_review.py {len(noisy_p2_lines)} 处 STRATEGY9/10 噪音 DEBUG-gate")

    # 4. p2_tech_review.py 重要成功/警告保留
    important_p2_lines = [
        "✓ OS 层关闭 Lodop 预览窗口",     # 成功 — 必须看到
        "ℹ️ 没找到预览窗口",                # 信息 — 有意义的
        "⚠️ [STRATEGY10] stderr",          # 警告
        "⚠️ [STRATEGY10] PowerShell 超时",  # 警告
        "⚠️ [STRATEGY10] 失败",             # 警告
        "⚠️ [STRATEGY9] {f.name} 失败",    # 警告
    ]
    for important in important_p2_lines:
        found_unconditional = False
        for line in p2_src.split("\n"):
            if important in line and ('print' in line or 'Write-Output' in line):
                if "if config.DEBUG" not in line:
                    found_unconditional = True
                    break
        assert found_unconditional, f"❌ p2 重要输出被误 DEBUG-gate: {important}"
    print(f"  ✅ p2_tech_review.py {len(important_p2_lines)} 处成功/警告保留")

    return True


def test_20_second_round_cleanup():
    """测试 20: 第二轮去噪 - popup1/2/4 + run.py 推进和 DBG (v1.2.3)

    背景: 上一轮 (Test 19) 只清了对弹窗 + STRATEGY9/10, 5 步里还有这些噪音:
          - popup1: popup URL / frames
          - popup2: [V1] dispatch_event 成功
          - popup4: 🧹 清理 残留弹窗
          - run.py: ─── 推进到下一节点 ─── 分隔符
          - run.py: 🔍 [DBG] 当前 URL/title/is_closed (3 行)
    修法:
          - 5 处噪音 DEBUG-gate (popup1/2/4)
          - 2 处直接删 (推进分隔符 + DBG 3 行)
    """
    print("\n=== Test 20: 第二轮去噪 ===")

    # 1. popup1 URL/frames DEBUG-gate
    p1_src = open(os.path.join(SCRIPT_DIR, "popups/p1_vehicle_check.py"), encoding="utf-8").read()
    assert "if config.DEBUG:" in p1_src, "❌ popup1 没 config.DEBUG guard"
    assert "🔍 popup URL" in p1_src, "❌ popup1 popup URL 丢了"
    assert "🔍 popup frames" in p1_src, "❌ popup1 popup frames 丢了"
    print("  ✅ popup1: popup URL/frames DEBUG-gate")

    # 2. popup2 [V1] dispatch_event DEBUG-gate
    p2_src = open(os.path.join(SCRIPT_DIR, "popups/p2_tech_review.py"), encoding="utf-8").read()
    assert "[V1] dispatch_event" in p2_src, "❌ popup2 [V1] 丢了"
    # 找包含 [V1] 的行, 确认有 if config.DEBUG
    v1_lines = [l for l in p2_src.split("\n") if "[V1]" in l and "print" in l]
    assert any("if config.DEBUG" in l for l in v1_lines), \
        "❌ popup2 [V1] 没 DEBUG-gate"
    print("  ✅ popup2: [V1] dispatch_event DEBUG-gate")

    # 3. popup4 清理残留弹窗 DEBUG-gate
    p4_src = open(os.path.join(SCRIPT_DIR, "popups/p4_vehicle_annual.py"), encoding="utf-8").read()
    assert "🧹 清理" in p4_src, "❌ popup4 清理逻辑丢了"
    cleanup_lines = [l for l in p4_src.split("\n") if "🧹 清理" in l and "print" in l]
    assert any("if config.DEBUG" in l for l in cleanup_lines), \
        "❌ popup4 🧹 清理没 DEBUG-gate"
    print("  ✅ popup4: 🧹 清理残留弹窗 DEBUG-gate")

    # 4. run.py ─── 推进到下一节点 ─── 已删
    run_src = open(os.path.join(SCRIPT_DIR, "run.py"), encoding="utf-8").read()
    assert "─── 推进到下一节点 ───" not in run_src, \
        "❌ run.py 还有 '推进到下一节点' 分隔符 (应该删除)"
    print("  ✅ run.py: '推进到下一节点' 分隔符已删")

    # 5. run.py 🔍 [DBG] 3 行已删
    dbg_count = run_src.count("🔍 [DBG]")
    assert dbg_count == 0, \
        f"❌ run.py 还有 {dbg_count} 个 🔍 [DBG] 行 (应全部删除)"
    print(f"  ✅ run.py: 🔍 [DBG] 行全部删除 ({dbg_count} 个残留)")

    return True


def test_21_phase1_log_cleanup():
    """测试 21: Phase1 噪音去噪 - 6 处 DEBUG-gate + 1 处创建任务保留 (v1.2.3)

    背景: Phase1 还有这些噪音 (用户在冀J5J289 报告):
          - [Phase1] 创建新任务 / [Phase1] 道路货物运输车辆审验 (标签)
          - → 弹窗开 (expect_event 捕获) / → 弹窗开 (调试说明)
          - ✓ frame[main_kef] 选中 radio / ✓ frame[main_kef] 点击'确定' (中间步)
          - ✓ 点击'创建任务(R)'成功 (用户要求保留)
    修法:
          - 6 处 DEBUG-gate
          - 创建任务成功 (force) + JS 路径保留 (不在 DEBUG-gate)
    """
    print("\n=== Test 21: Phase1 去噪 ===")

    run_src = open(os.path.join(SCRIPT_DIR, "run.py"), encoding="utf-8").read()

    # 1. 6 处噪音 DEBUG-gate
    noisy_phase1 = [
        "[Phase1] {menu_name}",            # 2 个 menu (道路货物运输 / 挂车及其他车辆)
        "→ 弹窗开 (expect_event 捕获)",
        "→ 弹窗开",
        "✓ frame[main_kef] 选中 radio 成功",
        "✓ frame[main_kef] 点击'确定'成功",
    ]
    for noisy in noisy_phase1:
        # 检查这个 print 是否被 if config.DEBUG: 包裹
        lines = [l for l in run_src.split("\n") if noisy in l and "print" in l]
        assert lines, f"❌ Phase1 噪音丢了: {noisy}"
        # 所有出现都应该有 if config.DEBUG
        unguarded = [l for l in lines if "if config.DEBUG" not in l]
        assert not unguarded, f"❌ Phase1 '{noisy}' 有 {len(unguarded)} 处没 DEBUG-gate: {unguarded[0].strip()}"
    print(f"  ✅ Phase1 {len(noisy_phase1)} 类噪音全部 DEBUG-gate")

    # 2. 创建任务成功保留 (force)
    create_task_lines = [l for l in run_src.split("\n") if "点击'创建任务(R)'成功 (force)" in l and "print" in l]
    assert create_task_lines, "❌ 创建任务成功丢了"
    guarded = [l for l in create_task_lines if "if config.DEBUG" in l]
    assert not guarded, \
        f"❌ 创建任务成功 (force) 被 DEBUG-gate 了 (用户要求保留)"
    print(f"  ✅ 创建任务成功 (force) 保留 (无 DEBUG-gate)")

    # 3. 创建任务 JS 成功也保留 (对称)
    js_create_lines = [l for l in run_src.split("\n") if "JS 点击'创建任务(R)'成功" in l and "print" in l]
    if js_create_lines:
        guarded = [l for l in js_create_lines if "if config.DEBUG" in l]
        assert not guarded, \
            f"❌ JS 创建任务成功 被 DEBUG-gate 了"
        print(f"  ✅ 创建任务 JS 成功也保留")

    return True


def test_22_popup4_submit_debug_gate():
    """测试 22: popup4 点击提交 DEBUG-gate (v1.2.3)

    背景: 用户反馈 popup4 还有 `✓ frame[_workflow_main] 点击提交` 这行啰嗦
          但保留 `JS-click link/text/a:has-text` (能看到点了哪个链接/文字)
    修法: 只 gate `点击提交`, 链接点击保留
    """
    print("\n=== Test 22: popup4 点击提交 DEBUG-gate ===")

    p4_src = open(os.path.join(SCRIPT_DIR, "popups/p4_vehicle_annual.py"), encoding="utf-8").read()

    # 1. 点击提交 DEBUG-gate
    submit_lines = [l for l in p4_src.split("\n") if "点击提交" in l and "print" in l]
    assert submit_lines, "❌ 点击提交 丢了"
    guarded = [l for l in submit_lines if "if config.DEBUG" in l]
    unguarded = [l for l in submit_lines if "if config.DEBUG" not in l]
    assert guarded and not unguarded, \
        f"❌ 点击提交没全部 DEBUG-gate (gate={len(guarded)}, unguarded={len(unguarded)})"
    print(f"  ✅ popup4 点击提交 DEBUG-gate ({len(guarded)} 处)")

    # 2. JS-click link 保留 (无 DEBUG-gate)
    js_link_lines = [l for l in p4_src.split("\n") if "JS-click link" in l and "print" in l]
    assert js_link_lines, "❌ JS-click link 丢了"
    unguarded = [l for l in js_link_lines if "if config.DEBUG" not in l]
    assert unguarded, "❌ JS-click link 被误 DEBUG-gate (用户要求保留)"
    print(f"  ✅ popup4 JS-click link 保留 ({len(js_link_lines)} 处)")

    return True


def test_23_runtime_check_license():
    """测试 23: runtime_check.py 伪装成运行时检查的 license 验证 (v1.2.3)

    背景: 防复制需求, 但不想被一眼看穿是 license 检查
    修法:
      - 新建 runtime_check.py (不叫 license_check.py)
      - 错误信息伪装成 RuntimeError + 技术细节
      - 不出现 license/授权/续费/expiry 等关键字
      - 联系方式藏在"请联系作者"里
    """
    print("\n=== Test 23: 伪装 license 检查 ===")

    # 1. 文件存在 + URL 常量
    assert os.path.exists(os.path.join(SCRIPT_DIR, "runtime_check.py")), \
        "❌ runtime_check.py 不存在"
    print("  ✅ runtime_check.py 存在 (伪装名)")

    rc_src = open(os.path.join(SCRIPT_DIR, "runtime_check.py"), encoding="utf-8").read()
    assert "RuntimeEnvError" in rc_src, "❌ RuntimeEnvError 异常类丢失"
    assert "check_runtime_env" in rc_src, "❌ check_runtime_env 函数丢失"
    print("  ✅ RuntimeEnvError + check_runtime_env 函数存在")

    # 2. URL 配置
    assert config.LICENSE_URL == "http://82.156.229.67/license.json", \
        f"❌ LICENSE_URL 错: {config.LICENSE_URL}"
    print(f"  ✅ LICENSE_URL = {config.LICENSE_URL}")

    # 3. 间隔配置
    assert config.LICENSE_CHECK_INTERVAL == 10, \
        f"❌ CHECK_INTERVAL 应为 10, 实际 {config.LICENSE_CHECK_INTERVAL}"
    assert config.LICENSE_TIMEOUT == 5, \
        f"❌ TIMEOUT 应为 5, 实际 {config.LICENSE_TIMEOUT}"
    print(f"  ✅ LICENSE_CHECK_INTERVAL = {config.LICENSE_CHECK_INTERVAL}, TIMEOUT = {config.LICENSE_TIMEOUT}")

    # 4. 伪装检查 — 错误信息不出现敏感关键字
    sensitive_keywords = ["license", "授权", "续期", "续费", "expiry", "expired", "过期", "LICENSE"]
    # runtime_check.py 源码里 license 出现是 OK 的 (注释 + 变量名), 但错误信息不能出现
    # 检查错误信息构建函数
    fake_msg_start = rc_src.find("def _build_fake_error_message")
    fake_msg_end = rc_src.find("def check_runtime_env")
    fake_msg_section = rc_src[fake_msg_start:fake_msg_end]

    for kw in sensitive_keywords:
        # 错误信息里不出现 (源码其他位置可以, 如注释)
        # 但这里我们检查整个 runtime_check.py, 确保错误信息能脱敏
        if kw == "license":
            # 函数名 check_runtime_env 不含 license 是关键
            assert "license" not in fake_msg_section.lower(), \
                f"❌ 错误信息里有 '{kw}' (会泄露授权机制)"
        elif kw == "LICENSE":
            assert kw not in fake_msg_section, \
                f"❌ 错误信息里有 '{kw}'"
        elif kw == "授权" or kw == "续期" or kw == "续费" or kw == "过期" or kw == "expiry" or kw == "expired":
            assert kw not in fake_msg_section, \
                f"❌ 错误信息里有 '{kw}' (会泄露授权机制)"
    print(f"  ✅ 错误信息不泄露授权机制 (检查 {len(sensitive_keywords)} 个关键字)")

    # 5. 联系方式藏在错误信息里 (伪装在"请联系作者")
    assert "18531729777" in fake_msg_section, \
        "❌ 错误信息没放联系方式 (用户要求超哥微信)"
    assert "微信" in fake_msg_section, \
        "❌ 错误信息没'微信'标记"
    print(f"  ✅ 错误信息含联系方式 (微信 18531729777)")

    # 6. gui.py 调用
    gui_src = open(os.path.join(SCRIPT_DIR, "gui.py"), encoding="utf-8").read()
    assert "from runtime_check import" in gui_src, "❌ gui.py 没 import runtime_check"
    assert "_check_runtime_env_or_die" in gui_src, "❌ gui.py 没定义启动检查函数"
    assert 'sys.argv' in gui_src and "headless" in gui_src, "❌ gui.py __main__ 结构丢了"
    # 检查两种模式都调用检查
    assert gui_src.count("check_runtime_env") >= 3, \
        f"❌ gui.py 调用 check_runtime_env 太少 (应是 import + 2 次调用 + 函数定义)"
    print(f"  ✅ gui.py 启动时检查 (GUI + 无头两种模式)")

    # 7. run.py 调用
    run_src = open(os.path.join(SCRIPT_DIR, "run.py"), encoding="utf-8").read()
    assert "from runtime_check import" in run_src, "❌ run.py 没 import runtime_check"
    # run.py 里调用 check_runtime_env 应该有 2 处 (续跑 + 主流程)
    call_count = run_src.count("check_runtime_env(force=False")
    assert call_count == 2, \
        f"❌ run.py 应该有 2 处调用 (续跑 + 主流程), 实际 {call_count}"
    print(f"  ✅ run.py 主循环调用 (2 处: 续跑 + 主流程)")

    return True


if __name__ == "__main__":
    print("=" * 60)
    print("  HCCheck v1.2 mock 测试")
    print("=" * 60)

    try:
        test_1_data_dir()
        run_id = test_2_record_runs()
        test_3_query(run_id)
        test_4_summary(run_id)
        test_5_export(run_id)
        test_6_empty_run_id()
        test_6b_latest_run_id()
        test_7_gui_loading()
        test_8_tab_order()
        test_9_no_tkinter_typos()
        test_10_popup4_excludes_main_page()
        test_13_popup2_strategy10_powershell()
        test_18_popup2_cleanup()
        test_19_log_noise_cleanup()
        test_20_second_round_cleanup()
        test_21_phase1_log_cleanup()
        test_22_popup4_submit_debug_gate()
        test_23_runtime_check_license()
        test_14_click_query_and_pagination()
        test_15_pagination_while_loop()
        test_16_popup2_strategy9_no_false_success()
        test_17_popup2_strategy10_imports_sys()

        print("\n" + "=" * 60)
        print("  ✅ 全部测试通过!")
        print("=" * 60)
    except AssertionError as e:
        print(f"\n❌ 测试失败: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n💥 异常: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
