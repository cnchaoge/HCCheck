"""运管站货车审验 - 数据库查询 / 导出模块

为 GUI 历史记录 Tab 提供只读 API:
- get_runs_by_run_id(run_id): 查询本次 run 的所有车 (按 start_time 升序)
- export_run_to_xlsx(run_id, output_path): 导出本次 run 为 xlsx (兼容旧格式)

xlsx 格式跟 v1.1 旧的 export_results_excel() 保持一致:
  车牌 | 类型 | 耗时(秒) | 状态 | 错误 | 开始时间

数据写入由 db.py 的 record_run() 负责,本模块只读不写。
"""
import sqlite3
import time as _time
from typing import List, Dict, Optional

import config


def get_runs_by_run_id(run_id: str) -> List[Dict]:
    """查询指定 run_id 的所有车辆处理结果

    Args:
        run_id: db.get_run_id() 返回的 UUID (一次运行一个)

    Returns:
        [{"plate": "冀J0E139", "flow_type": "带挂", "duration": 32.5,
          "status": "成功", "error": "", "start_time": 1234567890.0,
          "end_time": 1234567922.5}, ...]
        按 start_time 升序排列 (跑车的先后顺序)

        查询失败返回空列表 (不抛异常, GUI 不会崩)
    """
    if not run_id:
        return []
    try:
        with sqlite3.connect(config.DB_FILE, timeout=5) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute("""
                SELECT plate, flow_type, duration, status, error, start_time, end_time
                FROM runs
                WHERE run_id = ?
                ORDER BY start_time ASC
            """, (run_id,)).fetchall()
        return [dict(r) for r in rows]
    except Exception as e:
        if config.DEBUG:
            print(f"  调试 - get_runs_by_run_id 失败: {e}")
        return []


def get_run_summary(run_id: str) -> Dict:
    """统计本次 run 的汇总数据 (顶部状态栏用)

    Returns:
        {"total": 5, "success": 3, "failed": 2, "avg_duration": 28.4}
    """
    if not run_id:
        return {"total": 0, "success": 0, "failed": 0, "avg_duration": 0.0}
    try:
        with sqlite3.connect(config.DB_FILE, timeout=5) as conn:
            row = conn.execute("""
                SELECT
                    COUNT(*) AS total,
                    SUM(CASE WHEN status = '成功' THEN 1 ELSE 0 END) AS success,
                    SUM(CASE WHEN status = '失败' THEN 1 ELSE 0 END) AS failed,
                    AVG(duration) AS avg_duration
                FROM runs
                WHERE run_id = ?
            """, (run_id,)).fetchone()
        if row is None:
            return {"total": 0, "success": 0, "failed": 0, "avg_duration": 0.0}
        return {
            "total": row[0] or 0,
            "success": row[1] or 0,
            "failed": row[2] or 0,
            "avg_duration": round(row[3] or 0.0, 1),
        }
    except Exception as e:
        if config.DEBUG:
            print(f"  调试 - get_run_summary 失败: {e}")
        return {"total": 0, "success": 0, "failed": 0, "avg_duration": 0.0}


def export_run_to_xlsx(run_id: str, output_path: str) -> bool:
    """导出指定 run 的所有车为 xlsx 文件 (兼容旧格式)

    旧格式列: 车牌 | 类型 | 耗时(秒) | 状态 | 错误 | 开始时间

    Args:
        run_id: 本次运行的 UUID
        output_path: 输出 xlsx 完整路径 (含 .xlsx 后缀)

    Returns:
        True = 成功, False = 失败 (会打印原因)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  ❌ 缺少 openpyxl 依赖, 无法导出 xlsx")
        print("  💡 安装: pip install openpyxl")
        return False

    runs = get_runs_by_run_id(run_id)
    if not runs:
        print(f"  ⚠️ run_id={run_id} 没有数据, 不生成 xlsx")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "审验结果"

    # 表头
    headers = ["车牌", "类型", "耗时(秒)", "状态", "错误", "开始时间"]
    ws.append(headers)

    # 表头样式 (蓝底白字加粗, 居中)
    header_font = Font(bold=True, color="FFFFFF", name="Microsoft YaHei")
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    data_font = Font(name="Microsoft YaHei")
    for r in runs:
        start_time_str = _time.strftime(
            "%Y-%m-%d %H:%M:%S",
            _time.localtime(r["start_time"]) if r["start_time"] else _time.localtime()
        )
        duration = round(r["duration"], 1) if r["duration"] else 0.0
        ws.append([
            r["plate"] or "",
            r["flow_type"] or "",
            duration,
            r["status"] or "",
            r["error"] or "",
            start_time_str,
        ])
        # 数据行字体
        for cell in ws[ws.max_row]:
            cell.font = data_font

    # 列宽
    widths = {"A": 14, "B": 10, "C": 12, "D": 10, "E": 40, "F": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 冻结表头
    ws.freeze_panes = "A2"

    try:
        wb.save(output_path)
        print(f"  ✅ 已导出 {len(runs)} 条记录到: {output_path}")
        return True
    except Exception as e:
        print(f"  ❌ 保存 xlsx 失败: {e}")
        return False
