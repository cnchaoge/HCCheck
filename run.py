"""运管站货车审验自动化 - v3 (拆分 + 抽硬编码)

架构:
    run.py        - 主循环 + 续跑 + 分流
    config.py     - 所有可配置常量(站点改版只动这里)
    utils.py      - 通用工具(safe / pa / step / paste / screenshot)
    dialog.py     - "选下一处理人" 弹窗处理
    popups/       - 5 个 workflow popup 的处理函数

流程(手动验证通过):
    Select vehicle + create task
    不带挂:   popup1 车辆检测 → popup2 技术岗位审核 → popup3 业务岗位审核 → popup4 车辆年审 → popup5 归档
    带挂:     popup1 车辆检测 → popup2 业务岗位审核 → popup3 车辆年审1 → popup4 车辆年审2 → popup5 归档
"""
import config
from utils import safe, pa, step, paste_into, screenshot_on_error
from dialog import do_dialog
from popups import (
    handle_vehicle_check,
    handle_tech_review,
    handle_business_review,
    handle_vehicle_annual,
    handle_archive,
)
from playwright.sync_api import sync_playwright

# ========= 黑名单(失败 N 次的车牌不再自动重试) =========
SKIP_PLATES = set()

# ========= 处理结果记录（跑完后导出 Excel） =========
import time as _time
RESULTS = []  # [{plate, flow_type, start_time, end_time, status, error}]


def safe_input(prompt):
    """包装 input()，支持 stdin 被 GUI 关闭时优雅退出

    - 正常情况：调内置 input() 等用户回车
    - stdin 被关闭（GUI 强制停止时）：抛 SystemExit，让 run_main 退出
    """
    try:
        return input(prompt)
    except (EOFError, ValueError, KeyboardInterrupt):
        # stdin 被 GUI 关闭 / 用户 Ctrl+C → 查 FORCE_STOP
        if config.FORCE_STOP:
            raise SystemExit("用户强制停止")
        raise


def export_results_excel():
    """把处理结果导出为 Excel 报表"""
    if not RESULTS:
        return None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # 没装 openpyxl 就用 CSV
        import csv
        path = f"hccheck_results_{_time.strftime('%Y%m%d_%H%M%S')}.csv"
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=["车牌", "流程类型", "开始时间", "结束时间", "耗时(秒)", "状态", "错误信息"])
            writer.writeheader()
            for r in RESULTS:
                dur = ""
                if r["start_time"] and r["end_time"]:
                    dur = f"{r['end_time'] - r['start_time']:.1f}"
                writer.writerow({
                    "车牌": r["plate"],
                    "流程类型": r["flow_type"],
                    "开始时间": _time.strftime("%Y-%m-%d %H:%M:%S", _time.localtime(r["start_time"])) if r["start_time"] else "",
                    "结束时间": _time.strftime("%Y-%m-%d %H:%M:%S", _time.localtime(r["end_time"])) if r["end_time"] else "",
                    "耗时(秒)": dur,
                    "状态": r["status"],
                    "错误信息": r.get("error", ""),
                })
        return path

    wb = Workbook()
    ws = wb.active
    ws.title = "处理结果"
    headers = ["车牌", "流程类型", "开始时间", "结束时间", "耗时(秒)", "状态", "错误信息"]
    ws.append(headers)
    # 表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 数据
    for r in RESULTS:
        dur = ""
        if r["start_time"] and r["end_time"]:
            dur = round(r["end_time"] - r["start_time"], 1)
        ws.append([
            r["plate"],
            r["flow_type"],
            _time.strftime("%Y-%m-%d %H:%M:%S", _time.localtime(r["start_time"])) if r["start_time"] else "",
            _time.strftime("%Y-%m-%d %H:%M:%S", _time.localtime(r["end_time"])) if r["end_time"] else "",
            dur,
            r["status"],
            r.get("error", ""),
        ])
        # 状态列上色
        status_cell = ws.cell(row=ws.max_row, column=6)
        if r["status"] == "成功":
            status_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    # 列宽
    col_widths = [12, 10, 20, 20, 10, 8, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    path = f"hccheck_results_{_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path)
    return path


# ========= Helper 函数 =========
def goto_workbench(page):
    """切到工作台, 失败不抛异常（容错）"""
    contents = get_contents(page)
    try:
        safe(contents.get_by_role("link", name=config.MENU_WORKBENCH), timeout=5000).click()
        pa(2)
    except:
        pass


def get_main_kef(page):
    """获取主表格 frame locator"""
    return page.frame_locator(config.SELECTOR_FRAME_MAIN_KEF)


def get_contents(page):
    """获取左侧导航树 frame locator"""
    return page.frame_locator(config.SELECTOR_FRAME_CONTENTS)


# ========= 登录 + 导航 =========
def wait_for_login_and_navigate(page, username="", password="", auto_submit=False):
    page.goto(config.URL, wait_until="domcontentloaded")
    pa(1)

    print("=" * 50)
    print("  运管站货车审验自动化系统")
    print("=" * 50)
    print("\n[1/2] 登录系统")

    # 如果提供了账号密码，自动填入
    if username and password:
        print("  · 自动填入账号密码...")
        try:
            # 尝试常见的用户名/密码输入框选择器
            user_selectors = [
                "input[name='username']",
                "input[name='userName']",
                "input[name='loginName']",
                "input[name='j_username']",
                "input[type='text'][name*='user']",
                "input[type='text'][name*='name']",
                "#username",
                "#loginName",
            ]
            pass_selectors = [
                "input[name='password']",
                "input[name='passWord']",
                "input[name='j_password']",
                "input[type='password']",
                "#password",
            ]
            login_selectors = [
                "input[type='submit']",
                "button[type='submit']",
                "input[value='登录']",
                "input[value='登 录']",
                "button:has-text('登录')",
                "a:has-text('登录')",
                "#loginBtn",
                ".login-btn",
            ]
            # 填用户名
            user_filled = False
            for sel in user_selectors:
                try:
                    el = page.locator(sel).first
                    if el.is_visible(timeout=2000):
                        el.click()
                        el.fill(username)
                        user_filled = True
                        print(f"  ✓ 用户名已填入 (selector: {sel})")
                        break
                except:
                    continue
            if not user_filled:
                print("  ⚠️ 未找到用户名输入框，请手动登录")
                safe_input("  >>> 登录完成后按回车键继续... ")
                return

            pa(0.5)

            # 填密码
            pass_filled = False
            for sel in pass_selectors:
                try:
                    el = page.locator(sel).first
                    if el.is_visible(timeout=2000):
                        el.click()
                        el.fill(password)
                        pass_filled = True
                        print(f"  ✓ 密码已填入 (selector: {sel})")
                        break
                except:
                    continue
            if not pass_filled:
                print("  ⚠️ 未找到密码输入框，请手动登录")
                safe_input("  >>> 登录完成后按回车键继续... ")
                return

            pa(0.5)

            # 点登录按钮
            if auto_submit:
                for sel in login_selectors:
                    try:
                        el = page.locator(sel).first
                        if el.is_visible(timeout=2000):
                            el.click()
                            print(f"  ✓ 已点击登录按钮 (selector: {sel})")
                            pa(3)
                            break
                    except:
                        continue
                print("  · 等待登录完成...")
                pa(3)
            else:
                print("  · 账号密码已填入，等待手动点登录...")
                safe_input("  >>> 登录完成后按回车键继续... ")
        except Exception as e:
            print(f"  ⚠️ 自动登录失败: {e}")
            safe_input("  >>> 请手动登录后按回车键继续... ")
    else:
        print("  · 请手动登录")
        print("  · 登录后请确保左侧导航树已展开")
        safe_input("\n  >>> 登录完成后按回车键继续... ")

    contents = get_contents(page)
    # 导航树选择器备选(系统偶尔换 ID)
    nav_el = None
    for sel in config.NAV_TREE_SELECTORS:
        try:
            nav_el = contents.locator(sel).first
            nav_el.wait_for(state="visible", timeout=5000)
            break
        except:
            continue
    if nav_el is None:
        print("\n  ⚠️ 未检测到导航树,请手动展开左侧菜单")
        safe_input("  >>> 展开到[普货审验]页面后按回车... ")
    else:
        print("✅ 导航树已就绪")

    print("\n[2/2] 导航完成 ✅")

    # 登录后点击货运管理的展开按钮
    try:
        freight_row = contents.locator(f"a:has-text('{config.MENU_FREIGHT_MANAGE}')").locator("..").locator("img")
        if freight_row.count() > 0:
            freight_row.first.click()
            print("✅ 已点击 [货运管理] 展开按钮")
        else:
            contents.get_by_role("link", name=config.MENU_FREIGHT_MANAGE).click()
            print("✅ 已点击 [货运管理] 文字")
        pa(0.5)
    except Exception as e:
        print(f"⚠️ [货运管理] 点击失败: {e}")

    # 再点击工作台
    try:
        safe(contents.get_by_role("link", name=config.MENU_WORKBENCH), timeout=3000).click()
        print("✅ 已切换到 [工作台]")
        pa(1)
    except:
        print("⚠️ [工作台] 点击失败")

    step("开始批量处理")

# ========= Phase1 =========
def _check_workbench_for_task(page, plate):
    """检查工作台里这辆车有没有未完成的任务,有就接管
    Returns: (found, popup, step_name) or (False, None, None)
    """
    existing = _get_existing_popup(page)
    if existing:
        step_name = _detect_popup_step(existing, page)
        print(f"  发现已打开的弹窗 (步骤: {step_name})")
        return True, existing, step_name

    main_kef = get_main_kef(page)
    try:
        # 找表格里的任务行
        rows = main_kef.locator("table tbody tr")
        for r in range(rows.count()):
            cells = rows.nth(r).locator("td")
            for c in range(cells.count()):
                txt = cells.nth(c).text_content().strip()
                if txt == plate or txt == f"{plate} ":
                    row_text = " ".join([cells.nth(j).text_content().strip() for j in range(cells.count())])
                    step_name = _step_from_text(row_text)
                    print(f"  发现已有任务: {txt} (节点: {step_name})")

                    # 先检查弹窗是否已经打开
                    existing = _get_existing_popup(page)
                    if existing:
                        step_name = _detect_popup_step(existing, page)
                        print(f"  → 弹窗已存在,直接接管 (步骤: {step_name})")
                        return True, existing, step_name

                    # 点任务链接,尝试捕获新窗口
                    try:
                        with page.context.expect_event("page", timeout=15000) as page_info:
                            cells.nth(c).locator("a").first.click()
                        popup = page_info.value
                        print(f"  → 捕获到新窗口")
                        pa(2)
                        return True, popup, step_name
                    except:
                        # 超时: 可能弹窗没响应,再查一次
                        pa(3)
                        existing = _get_existing_popup(page)
                        if existing:
                            print(f"  → 弹窗已存在(延迟检测)")
                            return True, existing, step_name
                        print(f"  ⚠️ 点击任务但弹窗未出现,手动确认")
                        safe_input("  >>> 手动点开车牌链接后按回车...")
                        existing = _get_existing_popup(page)
                        if existing:
                            return True, existing, step_name
                        return False, None, None
    except Exception as e:
        print(f"  ⚠️ 检查工作台失败: {e}")
    return False, None, None


def _step_from_text(text):
    """从工作台行内文字识别当前步骤"""
    if config.NODE_TECH_REVIEW in text:
        return config.STEP_TECH_REVIEW
    if config.NODE_VEHICLE_CHECK in text:
        return config.STEP_VEHICLE_CHECK
    if config.NODE_BUSINESS_REVIEW in text:
        return config.STEP_BUSINESS_REVIEW
    if config.NODE_VEHICLE_ANNUAL in text:
        return config.STEP_VEHICLE_ANNUAL
    if config.NODE_ARCHIVE in text:
        return config.STEP_ARCHIVE
    return "unknown"


def read_workbench_node(page, plate):
    """从工作台读某辆车的当前节点
    Returns: STEP_XXX 或 None (未找到)
    """
    goto_workbench(page)
    main_kef = get_main_kef(page)
    try:
        rows = main_kef.locator("table tbody tr")
        for r in range(rows.count()):
            cells = rows.nth(r).locator("td")
            row_texts = [cells.nth(j).text_content().strip() for j in range(cells.count())]
            row_full = " ".join(row_texts)
            if plate in row_texts or f"{plate} " in row_texts:
                return _step_from_text(row_full)
    except:
        pass
    return None


def open_task_popup(page, plate):
    """在工作台点指定车牌的链接,返回新 popup
    Returns: 新 popup Page, 或 None
    """
    goto_workbench(page)
    main_kef = get_main_kef(page)
    try:
        links = main_kef.locator("a")
        for i in range(links.count() - 1, -1, -1):
            txt = links.nth(i).text_content().strip()
            if txt == plate or txt == f"{plate} ":
                try:
                    with page.context.expect_event("page", timeout=10000) as page_info:
                        links.nth(i).click(force=True)
                    return page_info.value
                except:
                    pa(3)
                    existing = _get_existing_popup(page)
                    if existing:
                        return existing
                    return None
    except Exception as e:
        print(f"  ⚠️ open_task_popup 异常: {e}")
    return None


def _phase1_select_and_create(page, plate, menu_name):
    """Phase1: 选车 + 创建任务
    先检查工作台有没有已有任务(续跑),有就接管
    没有就创建新任务,返回 (popup, STEP_VEHICLE_CHECK)
    """
    print(f"  [Phase1] {menu_name}")
    contents = get_contents(page)
    main_kef = get_main_kef(page)

    # 先去工作台
    safe(contents.get_by_role("link", name=config.MENU_WORKBENCH)).click()
    pa(2)

    # 先检查工作台有没有已有任务
    step("工作台检查")
    has_task, existing_popup, step_name = _check_workbench_for_task(page, plate)
    if has_task and existing_popup:
        print(f"   续跑起点: {step_name}")
        return existing_popup, step_name

    print("  工作台无已有任务,创建新任务")
    try:
        safe(main_kef.get_by_role("link", name=config.BTN_ADD), timeout=3000).click()
    except:
        pass
    pa(1)

    # 点菜单 → 用 expect_event 捕获新窗口
    pages_before = len(page.context.pages)
    try:
        with page.context.expect_event("page", timeout=15000) as page_info:
            safe(main_kef.get_by_role("link", name=menu_name)).click()
        popup1 = page_info.value
        print(f"  → 弹窗开 (expect_event 捕获)")
    except:
        # expect_event 超时, fallback 忙等
        print("  → expect_event 超时, fallback 忙等...")
        safe(main_kef.get_by_role("link", name=menu_name)).click()
        pa(2)
        popup1 = None
        for _ in range(20):
            if len(page.context.pages) > pages_before:
                popup1 = page.context.pages[-1]
                break
            pa(0.5)
    if not popup1:
        raise Exception("Phase1: 菜单弹窗未出现")
    print(f"  → 弹窗开")
    pa(1.5)

    # 选车
    safe(popup1.get_by_role("link", name=config.BTN_SELECT)).click()
    pa(1.5)

    diag   = popup1.frame_locator(config.SELECTOR_IFRAME_SUBMIT_DIAGCL)
    find   = diag.frame_locator(config.SELECTOR_FRAME_FIND_FRAME_KF)
    main_dlg = diag.frame_locator(config.SELECTOR_FRAME_MAIN_KEF)

    inp = safe(find.locator(config.INPUT_LICENSE_PLATE))
    paste_into(inp, plate)
    pa(1)

    try:
        safe(find.locator(config.SELECTOR_FLOW_ID), timeout=3000).select_option(config.FLOW_TYPE_ID)
    except:
        pass

    for sel in [find.get_by_role("button", name=config.BTN_QUERY),
                find.locator(f"input[value='{config.BTN_QUERY}']")]:
        try:
            safe(sel, timeout=4000).first.click()
            break
        except:
            continue
    pa(1.5)

    # 选中查询结果 - 尝试多种方式
    pa(1)
    _selected = False
    # 方式0 (优先级最高): 直接用 popup1.frames 拿 main_kef 点 radio
    for f in popup1.frames:
        if f.name == "main_kef":
            try:
                radios = f.get_by_role("radio")
                if radios.count() > 0:
                    radios.first.check(force=True)
                    _selected = True
                    print(f"  ✓ frame[main_kef] 选中 radio 成功")
                    break
            except Exception as e:
                print(f"  调试 - frame[main_kef] radio check 失败: {e}")
                continue
    # 方式1: 在 main_dlg 里找 radio/checkbox
    if not _selected:
        for sel in [main_dlg.get_by_role("radio").first,
                    main_dlg.locator("input[type='checkbox']").first]:
            try:
                safe(sel, timeout=3000).check()
                _selected = True
                break
            except:
                continue
    # 方式2: 直接在 diag 框架内找
    if not _selected:
        for sel in [diag.get_by_role("radio").first,
                    diag.locator("input[type='checkbox']").first]:
            try:
                safe(sel, timeout=2000).check()
                _selected = True
                break
            except:
                continue
    # 方式3: 在 popup1 最外层找
    if not _selected:
        for sel in [popup1.get_by_role("radio").first,
                    popup1.locator("input[type='checkbox']").first]:
            try:
                safe(sel, timeout=2000).check()
                _selected = True
                break
            except:
                continue
    # 方式4: 直接点击表格第一行(很多老系统点行 = 选中)
    if not _selected:
        for frame in [main_dlg, diag, popup1]:
            try:
                row = frame.locator("table tr").first
                safe(row, timeout=2000).click()
                _selected = True
                break
            except:
                continue
    if not _selected:
        print("  ⚠️ 未找到可勾选的 checkbox,请手动勾选后按回车")
        safe_input(">>> 勾选后按回车继续...")
    pa(1)

    # 点击"确定"按钮 - 尝试多个 frame 位置 (容错)
    _ok_clicked = False
    # 直接用 Playwright 的 frame API 查找 (避免 frame_locator 的路径问题)
    for f in popup1.frames:
        if f.name == "main_kef":
            try:
                btn = f.get_by_role("button", name=config.BTN_OK)
                btn.wait_for(state="visible", timeout=5000)
                btn.click()
                _ok_clicked = True
                print(f"  ✓ frame[main_kef] 点击'确定'成功")
                break
            except Exception as e:
                print(f"  ⚠️ frame[main_kef] 点击失败: {e}")
                continue
    if not _ok_clicked:
        for frame in [main_dlg, diag, popup1]:
            try:
                safe(frame.get_by_role("button", name=config.BTN_OK), timeout=3000).click()
                _ok_clicked = True
                break
            except:
                continue
    if not _ok_clicked:
        print("  ⚠️ 未找到'确定'按钮,尝试文本定位...")
        try:
            main_dlg.get_by_text(config.BTN_OK).first.click()
            _ok_clicked = True
        except:
            # 终极兑底: 在 popup 主页面遍历所有 iframe 查找'确定'按钮
            print("  ⚠️ 遍历所有 iframe 查找'确定'按钮...")
            # 先调试: 列出所有 frame 和 button
            print(f"  调试 - popup1.frames: {[f.name for f in popup1.frames]}")
            for f in popup1.frames:
                try:
                    btns_info = f.evaluate("""() => {
                        const result = [];
                        const all = document.querySelectorAll('input[type=button],input[type=submit],button,a');
                        for (const b of all) {
                            const t = (b.value || b.textContent || '').trim();
                            if (t) result.push(t.substring(0, 20));
                        }
                        return result.join('|');
                    }""")
                    print(f"  调试 - frame[{f.name}] buttons: {btns_info}")
                except Exception as e:
                    print(f"  调试 - frame[{f.name}] 读取失败: {e}")
            try:
                ok = popup1.evaluate("""() => {
                    // 在主文档查找
                    let btns = document.querySelectorAll('input[type=button],button,a');
                    for (const b of btns) {
                        if (b.value === '确定' || b.textContent.trim() === '确定') {
                            b.click();
                            return 'main:' + (b.tagName);
                        }
                    }
                    // 遍历所有 iframe
                    const iframes = document.querySelectorAll('iframe,frame');
                    for (const f of iframes) {
                        try {
                            const doc = f.contentDocument || f.contentWindow.document;
                            const btns2 = doc.querySelectorAll('input[type=button],button,a');
                            for (const b of btns2) {
                                if (b.value === '确定' || b.textContent.trim() === '确定') {
                                    b.click();
                                    return 'iframe:' + (f.name || f.id);
                                }
                            }
                        } catch(e) {}
                    }
                    return null;
                }""")
                if ok:
                    _ok_clicked = True
                    print(f"  ✓ JS 全局查找点击'确定'成功 ({ok})")
                else:
                    print("  ❌ JS 全局查找仍未找到'确定'按钮")
            except Exception as e:
                print(f"  ❌ JS 全局查找异常: {e}")
            if not _ok_clicked:
                print("  ❌ 确定按钮点击失败")
    pa(1.5)
    try:
        safe(popup1.get_by_role("link", name=config.BTN_START_TASK), timeout=3000).click(force=True)
        print(f"  ✓ 点击'创建任务(R)'成功 (force)")
    except Exception as e:
        print(f"  ⚠️ force 点击失败, 尝试 JS 点击: {e}")
        try:
            popup1.get_by_role("link", name=config.BTN_START_TASK).first.evaluate("el => el.click()")
            print(f"  ✓ JS 点击'创建任务(R)'成功")
        except Exception as e2:
            print(f"  ❌ 创建任务点击失败: {e2}")
            raise
    pa(3)
    pa(2)

    # 调试: 打印弹窗状态
    print(f"  🔍 popup1 URL: {popup1.url[:80]}")
    print(f"  🔍 popup1 frames: {[f.name for f in popup1.frames if f.name]}")

    return popup1, config.STEP_VEHICLE_CHECK

# ========= 菜单展开辅助 =========
def _click_expand_freight_manage(contents):
    """多重策略点击货运管理展开按钮
    策略: 1. 点 img 图标  2. 点文字 + force  3. JS 强制触发
    """
    # 检查是否已展开 (看普货审验是否可见)
    try:
        contents.get_by_text(config.MENU_NORMAL_REVIEW).first.wait_for(state="visible", timeout=1000)
        print(f"  ✅ 菜单已展开")
        return True
    except:
        pass

    # 策略1: 点 img 图标
    try:
        freight_row = contents.locator(f"a:has-text('{config.MENU_FREIGHT_MANAGE}')").locator("..").locator("img")
        if freight_row.count() > 0:
            freight_row.first.click()
            pa(1)
            # 验证是否展开
            try:
                contents.get_by_text(config.MENU_NORMAL_REVIEW).first.wait_for(state="visible", timeout=2000)
                print(f"  ✅ 点图标展开成功")
                return True
            except:
                pass
    except:
        pass

    # 策略2: 点文字 (force=True)
    try:
        link = contents.get_by_role("link", name=config.MENU_FREIGHT_MANAGE)
        if link.count() > 0:
            link.first.click(force=True)
            pa(1)
            try:
                contents.get_by_text(config.MENU_NORMAL_REVIEW).first.wait_for(state="visible", timeout=2000)
                print(f"  ✅ 点文字展开成功")
                return True
            except:
                pass
    except:
        pass

    # 策略3: JS 强制点击文字
    try:
        contents.locator(f'a:has-text("{config.MENU_FREIGHT_MANAGE}")').first.evaluate("el => el.click()")
        pa(1)
        try:
            contents.get_by_text(config.MENU_NORMAL_REVIEW).first.wait_for(state="visible", timeout=2000)
            print(f"  ✅ JS 展开成功")
            return True
        except:
            pass
    except:
        pass

    # 策略4: 重试 (再点一次图标)
    try:
        freight_row = contents.locator(f"a:has-text('{config.MENU_FREIGHT_MANAGE}')").locator("..").locator("img")
        if freight_row.count() > 0:
            freight_row.first.click()
            pa(2)
    except:
        pass

    print(f"  ⚠️ 菜单展开可能未成功")
    return False


def _click_normal_review_link(contents):
    """多重策略点击普货审验链接"""
    # 策略1: role 定位
    try:
        safe(contents.get_by_role("link", name=config.MENU_NORMAL_REVIEW), timeout=5000).click()
        print(f"  ✓ 点普货审验 (role)")
        return True
    except:
        pass
    # 策略2: 文本定位
    try:
        contents.get_by_text(config.MENU_NORMAL_REVIEW).first.click()
        print(f"  ✓ 点普货审验 (text)")
        return True
    except:
        pass
    # 策略3: JS 强制点击
    try:
        contents.locator(f'a:has-text("{config.MENU_NORMAL_REVIEW}")').first.evaluate("el => el.click()")
        print(f"  ✓ 点普货审验 (JS)")
        return True
    except:
        pass
    print(f"  ⚠️ 普货审验点击失败")
    return False


def _verify_in_normal_review(page):
    """验证 main_kef 是否在普货审验列表页面
    特征: 包含'年审列表'或'年审审批流水号'或列头包含'车牌号'
    """
    main_kef = get_main_kef(page)
    try:
        text = main_kef.locator("body").text_content(timeout=3000)
        if any(keyword in text for keyword in ["年审列表", "年审审批流水号", "申请日期"]):
            return True
    except:
        pass
    return False

# ========= Phase2 通用 =========
def _get_existing_popup(page):
    for p in page.context.pages:
        if p == page: continue
        try:
            u = (p.url or "").lower()
            if "workflow" in u:
                pa(2)
                return p
        except:
            continue
    return None

def _wait_popup_with_hint(page, hint="请手动点车牌打开弹窗"):
    popup = _get_existing_popup(page)
    if popup:
        print("  → 接管已有popup")
        return popup

    # 先切到工作台(当前可能在年审列表)
    print("  → 切到工作台找新任务...")
    contents = get_contents(page)
    try:
        safe(contents.get_by_role("link", name=config.MENU_WORKBENCH), timeout=5000).click()
        pa(2)
    except:
        print("  工作台链接找不到,假设已在工作台")

    main_kef = get_main_kef(page)
    link_clicked = None
    try:
        links = main_kef.locator("a")
        print(f"  调试 - main_kef 链接数: {links.count()}")
        for i in range(links.count() - 1, -1, -1):
            txt = links.nth(i).text_content().strip()
            if config.PLATE_RE.match(txt):
                link_clicked = txt
                print(f"  调试 - 找到车牌链接: {txt}, 点击中...")
                # 用 expect_event 捕获新窗口
                try:
                    with page.context.expect_event("page", timeout=15000) as page_info:
                        links.nth(i).click(force=True)
                    popup = page_info.value
                    print(f"  → 捕获到新窗口: {link_clicked}")
                    pa(2)
                    return popup
                except Exception as e:
                    print(f"  ⚠️ expect_event 超时: {e}")
                    # 超时, fallback
                    pa(3)
                    existing = _get_existing_popup(page)
                    if existing:
                        print(f"  → 弹窗已存在(延迟检测)")
                        return existing
                # 找到一个车牌就尝试,不管成不成都 break 不再继续试下一个
                break
    except Exception as e:
        print(f"  ⚠️ 查找车牌链接异常: {e}")

    # 如果上面 break 了但没拿到 popup,fallback 请示用户
    print(f"  ⏳ 自动跳转失败, {hint}")
    if link_clicked:
        print(f"  ⏳ 已点车牌 {link_clicked} 但弹窗未出现")
    safe_input(">>> 按回车继续...")
    pa(5)  # 多等一会
    popup = _get_existing_popup(page)
    if popup:
        print("  → 接管手动打开的popup")
        return popup
    # 再查一次
    pa(3)
    popup = _get_existing_popup(page)
    if popup:
        return popup
    raise Exception("弹窗未出现")

def _close_all_popups(page):
    """关掉所有非主页面的弹窗,防止残留"""
    closed = 0
    for p in page.context.pages:
        if p != page:
            try:
                p.close()
                closed += 1
            except:
                pass
    if closed:
        print(f"  清理 {closed} 个残留弹窗")
    pa(1)

def _advance_to_next_popup(page, prev_popup, hint="请手动点车牌打开弹窗"):
    """等待下一个业务弹窗出现
    策略: 系统完成后会从过渡页打开新弹窗,需要用 expect_event 等待
    """
    # 主动点工作台任务链接 (这是系统跳转业务弹窗的常规方式)
    pa(2)
    try:
        contents = get_contents(page)
        safe(contents.get_by_role("link", name=config.MENU_WORKBENCH), timeout=5000).click()
        pa(2)
        print(f"  调试 - 已切到工作台")
    except Exception as e:
        print(f"  调试 - 切工作台失败: {e}")

    # 查现有 pages 里有没有新弹窗 (可能已被自动打开)
    for p in page.context.pages:
        if p != page and p != prev_popup:
            try:
                if "workflow" in (p.url or "").lower() and "Flag=workflow_fra" in (p.url or ""):
                    print(f"  调试 - 发现现有新弹窗: {p.url[:80]}")
                    return p
            except:
                continue

    # 用 expect_event 等待新弹窗 (最长 15 秒)
    try:
        with page.context.expect_event("page", timeout=15000) as page_info:
            pass  # 触发器:什么都不做,只是等新 page 打开
        new_popup = page_info.value
        print(f"  调试 - expect_event 捕获新弹窗: {new_popup.url[:80]}")
        return new_popup
    except Exception as e:
        print(f"  调试 - expect_event 超时: {e}, 尝试其他方式")

    # fallback: 在工作台里点车牌链接
    main_kef = get_main_kef(page)
    try:
        links = main_kef.locator("a")
        for i in range(links.count() - 1, -1, -1):
            txt = links.nth(i).text_content().strip()
            if config.PLATE_RE.match(txt):
                with page.context.expect_event("page", timeout=10000) as page_info:
                    links.nth(i).click(force=True)
                new_popup = page_info.value
                print(f"  调试 - 点击车牌链接捕获新弹窗: {txt}")
                return new_popup
    except Exception as e:
        print(f"  调试 - 点击车牌链接失败: {e}")

    # 最后 fallback: 调 _wait_popup_with_hint
    return _wait_popup_with_hint(page, hint)

# ========= 步骤识别 =========
def _detect_popup_step(popup, page=None):
    """检测弹窗当前步骤
    先查特征元素(最准确),再扫 body 文字
    """
    # 检查弹窗是否还活着,如果已关闭且给了 page,从 context 重新找
    def _get_alive_popup(p):
        try:
            _ = p.url  # 访问 url 如果页面已关闭会报错
            return p
        except:
            return None

    popup = _get_alive_popup(popup) or popup
    if popup is None or _get_alive_popup(popup) is None:
        # 弹窗已关闭,从 context 里重新找一个 workflow 页面
        if page is not None:
            for p in page.context.pages:
                if _get_alive_popup(p) is not None and "workflow" in (p.url or "").lower():
                    popup = p
                    print(f"  调试 - 接管新弹窗: {popup.url[:80]}")
                    break

    if popup is None or _get_alive_popup(popup) is None:
        # 弹窗已关闭, 从 context 里找活的 workflow 页面
        if page is not None:
            print(f"  调试 - popup 已关闭, 从 context.pages 找新 popup...")
            for p in page.context.pages:
                if _get_alive_popup(p) is not None and "workflow" in (p.url or "").lower():
                    popup = p
                    print(f"  调试 - 接管新 popup: {popup.url[:80]}")
                    break
        if popup is None or _get_alive_popup(popup) is None:
            # 主动调用 _advance_to_next_popup 获取下一个 popup
            print(f"  调试 - 调用 _advance_to_next_popup 获取下一个 popup...")
            popup = _advance_to_next_popup(page, popup)
            if popup is not None and _get_alive_popup(popup) is not None:
                print(f"  调试 - advance 拿到 popup: {popup.url[:80]}")
            else:
                print(f"  ⚠️ 无法找到可用弹窗")
                return "unknown"

    # 先等弹窗 URL 稳定 (避免拿到上一页面的残留)
    print(f"  调试 - 弹窗 URL: {popup.url[:80]}")
    pa(5)  # 多等 5 秒, 让系统跳转完成
    # 增加超时次数 (原 8 次 = 8 秒, 新 20 次 = 20 秒)
    for attempt in range(30):
        # 重新从 context.pages 拿最新的活的 workflow 页面 (避免过期引用)
        if page is not None:
            alive_workflows = [p for p in page.context.pages
                               if _get_alive_popup(p) is not None and "workflow" in (p.url or "").lower()]
            if alive_workflows:
                # 拿最后一个 (最新的) - 不管 url 是否变化都接管
                latest = alive_workflows[-1]
                if _get_alive_popup(latest) is not None:
                    if latest.url != popup.url:
                        popup = latest
                        print(f"  调试 - 重新接管弹窗 (attempt={attempt}): {popup.url[:80]}")
                    elif _get_alive_popup(popup) is None:
                        popup = latest
                        print(f"  调试 - 弹窗死了, 重接 (attempt={attempt}): {popup.url[:80]}")
        try:
            # 优先用 popup.frames 直接访问 _workflow_main (避免 frame_locator 路径问题)
            wf_frame = None
            for f in popup.frames:
                if f.name == "_workflow_main":
                    wf_frame = f
                    break
            if wf_frame is None:
                # fallback 用 frame_locator
                wf_frame = popup.frame_locator(config.SELECTOR_FRAME_WORKFLOW_MAIN)
            # 1 特征元素优先
            if wf_frame.locator(f"a:has-text('{config.DETECT_YEAR_CHECK}')").count() > 0:
                print(f"  调试 - 检测到'年度审验'特征 (attempt={attempt})")
                return config.STEP_VEHICLE_ANNUAL
            if wf_frame.locator(f"a:has-text('{config.DETECT_PRINT}')").count() > 0:
                print(f"  调试 - 检测到'打印'特征 (attempt={attempt})")
                return config.STEP_TECH_REVIEW
            if wf_frame.get_by_role("button", name=config.BTN_COMPLETE).count() > 0:
                print(f"  调试 - 检测到'完成'按钮 (attempt={attempt})")
                return config.STEP_ARCHIVE

            # 2 页面文字兜底
            body_text = wf_frame.locator("body").text_content(timeout=2000)
            if config.NODE_VEHICLE_ANNUAL in body_text and config.DETECT_YEAR_CHECK in body_text:
                print(f"  调试 - body 检测到'车辆年审'+'年度审验' (attempt={attempt})")
                return config.STEP_VEHICLE_ANNUAL
            if config.NODE_VEHICLE_CHECK in body_text:
                print(f"  调试 - body 检测到'车辆检测' (attempt={attempt})")
                return config.STEP_VEHICLE_CHECK
            if config.NODE_TECH_REVIEW in body_text:
                print(f"  调试 - body 检测到'技术岗位审核' (attempt={attempt})")
                return config.STEP_TECH_REVIEW
            if config.NODE_BUSINESS_REVIEW in body_text:
                print(f"  调试 - body 检测到'业务岗位审核' (attempt={attempt})")
                return config.STEP_BUSINESS_REVIEW
            if config.NODE_ARCHIVE in body_text:
                print(f"  调试 - body 检测到'归档' (attempt={attempt})")
                return config.STEP_ARCHIVE
        except Exception as e:
            if attempt < 3:
                print(f"  调试 - 检测异常 (attempt={attempt}): {e}")
        pa(1)
    print(f"  ⚠️ 检测超时, 返回 unknown")
    return "unknown"

# ========= 两条流程 =========
def process_unmarked(page, plate, run_from_step=None, processed=0):
    """不带挂:popup1 车辆检测 → popup2 技术 → popup3 业务 → popup4 年审 → popup5 归档
    工作台节点驱动: 每完成一个 popup 后重读工作台,决定下一步
    """
    context = page.context
    print(f"\n  🚗 不带挂: {plate}")

    # 首次创建任务 (如果有 run_from_step,说明任务已存在,跳过创建)
    if run_from_step is None:
        print("  [Phase1] 创建新任务")
        popup, _ = _phase1_select_and_create(page, plate, config.MENU_FOR_NORMAL)
        handle_vehicle_check(popup, plate)
        print("\n  ─── 推进到下一节点 ───")
    else:
        print(f"  续跑: 从 {run_from_step} 开始")

    # 按工作台节点循环驱动
    step_handlers = {
        config.STEP_VEHICLE_CHECK: lambda p: handle_vehicle_check(p, plate),
        config.STEP_TECH_REVIEW: lambda p: handle_tech_review(p, context, page, plate),
        config.STEP_BUSINESS_REVIEW: lambda p: handle_business_review(p, context, plate, action_type=config.ACTION_SUBMIT_VEHICLE_ANNUAL),
        config.STEP_VEHICLE_ANNUAL: lambda p: handle_vehicle_annual(p, context, plate, action_type=config.ACTION_SUBMIT_ARCHIVE, category=config.CATEGORY_ROLE),
        config.STEP_ARCHIVE: lambda p: handle_archive(p, page, plate),
    }

    max_steps = 10  # 防止死循环
    for _ in range(max_steps):
        # 读工作台当前节点
        current_node = read_workbench_node(page, plate)
        print(f"  📋 工作台节点: {current_node}")

        if current_node is None:
            print(f"  ⚠️ 工作台未找到 {plate}, 可能已完成")
            break
        if current_node not in step_handlers:
            # 可能是未识别节点或已完成
            print(f"  ⚠️ 未识别节点 {current_node}, 退出")
            break

        # 打开 popup
        popup = open_task_popup(page, plate)
        if popup is None:
            print(f"  ⚠️ 打开 popup 失败")
            break

        # 执行对应 handler
        step_names_zh = {
            config.STEP_VEHICLE_CHECK: "车辆检测",
            config.STEP_TECH_REVIEW: "技术岗位审核",
            config.STEP_BUSINESS_REVIEW: "业务岗位审核",
            config.STEP_VEHICLE_ANNUAL: "车辆年审",
            config.STEP_ARCHIVE: "归档",
        }
        step_name = step_names_zh.get(current_node, current_node)
        # 人性化提示: 带车牌 + 步骤说明
        print(f"\n  ════════════════════════════════════")
        print(f"  🚗 车牌: {plate}")
        print(f"  📋 当前步骤: {step_name} (第 {processed} 辆车)")
        print(f"  ⏳ 正在处理中...")
        print(f"  ════════════════════════════════════")
        step_handlers[current_node](popup)

        # 完成后系统会自动关闭 popup,下一轮循环会重新读工作台
        pa(2)
        # 如果是归档,下一轮循环会发现节点变成未识别,退出
        if current_node == config.STEP_ARCHIVE:
            print(f"  ✅ {plate} 全部完成!")
            return

    print(f"  ✅ {plate} 流程结束")

def process_marked(page, plate, run_from_step=None, processed=0):
    """带挂流程: 业务岗位审核 → 车辆年审 → 归档 (3 个 popup)
    工作台节点驱动: 每完成一个 popup 后重读工作台,决定下一步
    """
    context = page.context
    print(f"\n  🚛 带挂: {plate}")

    # 首次创建任务
    if run_from_step is None:
        print("  [Phase1] 创建新任务")
        popup, _ = _phase1_select_and_create(page, plate, config.MENU_FOR_TRAILER)
    else:
        print(f"  续跑: 从 {run_from_step} 开始")
        popup = None

    # 按工作台节点循环驱动
    # 带挂 popup1 (业务岗位审核)  →  动作类型=提交年审,  类别=角色
    # 带挂 popup2 (车辆年审)       →  动作类型=提交归档,  类别=发起人
    # 带挂 popup3 (归档)            →  点完成 + 2次确认
    step_handlers = {
        config.STEP_BUSINESS_REVIEW: lambda p: handle_business_review(p, context, plate, action_type=config.ACTION_SUBMIT_YEAR_CHECK),
        config.STEP_VEHICLE_ANNUAL: lambda p: handle_vehicle_annual(p, context, plate, action_type=config.ACTION_SUBMIT_ARCHIVE, category=config.CATEGORY_INITIATOR),
        config.STEP_ARCHIVE: lambda p: handle_archive(p, page, plate),
    }

    max_steps = 10  # 防止死循环
    for _ in range(max_steps):
        # 读工作台当前节点
        current_node = read_workbench_node(page, plate)
        print(f"  📋 工作台节点: {current_node}")

        if current_node is None:
            print(f"  ⚠️ 工作台未找到 {plate}, 可能已完成")
            break
        if current_node not in step_handlers:
            print(f"  ⚠️ 未识别节点 {current_node}, 退出")
            break

        # 打开 popup
        if popup is None:
            popup = open_task_popup(page, plate)
        if popup is None:
            print(f"  ⚠️ 打开 popup 失败")
            break

        # 执行对应 handler
        step_names = {
            config.STEP_BUSINESS_REVIEW: "业务岗位审核",
            config.STEP_VEHICLE_ANNUAL: "车辆年审",
            config.STEP_ARCHIVE: "归档",
        }
        print(f"\n  ═══════════════════")
        print(f"  📋 popup: {step_names.get(current_node, current_node)}")
        print(f"  ═══════════════════")
        step_handlers[current_node](popup)

        # 完成后系统会自动关闭 popup
        pa(2)
        popup = None  # 下轮循环重新开

        # 如果是归档,下一轮循环会发现节点变成未识别,退出
        if current_node == config.STEP_ARCHIVE:
            print(f"  ✅ {plate} 全部完成!")
            return

    print(f"  ✅ {plate} 流程结束")


def _handle_year_check_first(popup, context, plate):
    """带挂流程的"车辆年审1" - 跟 popup4 几乎一样,但提交后选"提交年审"(不是"提交归档")"""
    print("\n  ═══════════════════")
    print("  📋 带挂: 车辆年审1")
    print("  ═══════════════════")
    step("车辆年审1: 准备点击年度审验")
    wf = popup.frame_locator(config.SELECTOR_FRAME_WORKFLOW_MAIN)
    conn = wf.frame_locator(config.SELECTOR_IFRAME_IFRAME_CONTENT)

    # 点年度审验(3 种文字变体)
    for sel_text in config.YEAR_CHECK_TEXTS:
        try:
            el = conn.get_by_role("link", name=sel_text)
            if el.count() > 0:
                safe(el.first, timeout=3000).click()
                print(f"  ✓ 点击'{sel_text}'")
                break
        except:
            try:
                el = conn.locator(f'text="{sel_text}"')
                if el.count() > 0:
                    el.first.dispatch_event("click")
                    print(f"  ✓ 点击'{sel_text}'")
                    break
            except:
                continue
    else:
        print("  ⚠️ 年度审验找不到")
    pa(1.5)
    try:
        safe(wf.get_by_role("button", name=config.BTN_SUBMIT), timeout=5000).click()
        print("  ✓ 点击提交")
    except:
        print("  ⚠️ 提交按钮找不到")
    pa(2)
    do_dialog(popup,
              action_type=config.ACTION_SUBMIT_YEAR_CHECK,
              category=config.CATEGORY_ROLE)


def _phase2_finish(popup, page):
    # 等 dialog 监听器消化完归档后的 confirm 弹窗
    pa(1)
    try: popup.close()
    except: pass
    pa(1)
    contents = get_contents(page)
    try:
        safe(contents.get_by_role("link", name=config.MENU_NORMAL_REVIEW), timeout=8000).click()
    except:
        pass
    pa(2)

def dispatch(page, plate, run_from_step=None, processed=0):
    # 通知 GUI：当前正在处理 plate
    config.CURRENT_PLATE = plate
    try:
        if "挂" in plate:
            process_marked(page, plate, run_from_step, processed)
        else:
            process_unmarked(page, plate, run_from_step, processed)
    finally:
        # 无论成功/异常,都清空状态,让 GUI 知道进入空闲
        config.CURRENT_PLATE = ""

# ========= 从列表拿车牌 =========
def get_next_plate_from_list(page):
    """从年审列表拿第一个无流水号的车牌(FIFO),跳过黑名单
    扫描多个 table (年审列表可能分多个表格)
    Columns: Seq|Radio|ApprovalNo|AppNo|Name|Date|Phone|Plate|Color|VIN|Source
    ApprovalNo at index 2, Plate at index 7
    """
    main_kef = get_main_kef(page)
    # 扫描所有 table (可能分多页)
    try:
        tables = main_kef.locator("table")
        for t_idx in range(tables.count()):
            rows = tables.nth(t_idx).locator("tbody tr")
            for r in range(rows.count()):
                cells = rows.nth(r).locator("td")
                # 检查年审审批流水号列(索引2)是否为空
                if cells.count() > 2:
                    flow_no = cells.nth(2).text_content().strip()
                    if flow_no:
                        continue  # 有流水号,跳过
                # 找车牌号(扫描所有列)
                for c in range(cells.count()):
                    txt = cells.nth(c).text_content().strip()
                    if config.PLATE_RE.match(txt):
                        if txt in SKIP_PLATES:
                            print(f"  ⏭ 跳过已拉黑: {txt}")
                            continue
                        return txt
    except Exception as e:
        print(f"  ⚠️ 读取年审列表失败: {e}")
    # fallback: 扫全文找第一个未拉黑的车牌
    try:
        all_text = main_kef.locator("body").text_content()
        for m in config.PLATE_RE.findall(all_text):
            if m not in SKIP_PLATES:
                print(f"  全文搜索取到车牌: {m}")
                return m
    except Exception as e:
        print(f"  ⚠️ 全文搜索车牌失败: {e}")
    # 调试: 打印 main_kef 内容
    try:
        content = main_kef.locator("body").text_content(timeout=3000)[:300]
        print(f"  🔍 main_kef内容(前300字): {content}")
    except:
        pass
    return None

# ========= 主流程 =========
def _filter_page_error(err):
    """过滤 IE 专属 API 错误 (Chrome 不支持但页面在用,无害)"""
    msg = str(err)
    ie_apis = ["attachEvent", "createPopup", "showModalDialog"]
    if any(api in msg for api in ie_apis):
        return  # 静默忽略
    # 其他错误才显示
    print(f"  ⚠️ 页面错误: {msg}")


def main():
    with sync_playwright() as p:
        browser = p.chromium.launch(channel="chrome", headless=config.HEADLESS)
        context = browser.new_context(viewport={"width": 1600, "height": 900})
        page = context.new_page()
        page.on("pageerror", lambda e: _filter_page_error(e))
        # 自动处理所有对话框(如confirm、alert、prompt"),吃掉关闭后的异常
        def _auto_accept_dialog(dialog):
            try:
                dialog.accept()
            except:
                pass
        context.on("dialog", _auto_accept_dialog)

        try:
            wait_for_login_and_navigate(
                page,
                username=config.LOGIN_USERNAME,
                password=config.LOGIN_PASSWORD,
                auto_submit=config.LOGIN_AUTO_SUBMIT,
            )
            print("\n" + "=" * 55)
            print("  开始批量审验处理")
            print("=" * 55)

            empty_count = 0
            processed = 0
            failures = {}

            while True:
                # === 检查强制停止信号 ===
                if config.FORCE_STOP:
                    print("\n⛔ 强制停止信号已接收，立即退出")
                    break

                # === 第一步: 检查工作台是否有未完成的任务(续跑) ===
                goto_workbench(page)

                main_kef = get_main_kef(page)
                workbench_plates = []  # [(plate, current_node, flow_name), ...]
                try:
                    rows = main_kef.locator("table tbody tr")
                    for r in range(rows.count()):
                        cells = rows.nth(r).locator("td")
                        row_texts = [cells.nth(j).text_content().strip() for j in range(cells.count())]
                        row_full = " ".join(row_texts)
                        # 找车牌号
                        for t in row_texts:
                            if config.PLATE_RE.match(t) and t not in SKIP_PLATES:
                                current_node = ""
                                if config.NODE_TECH_REVIEW in row_full:
                                    current_node = config.NODE_TECH_REVIEW
                                elif config.NODE_VEHICLE_CHECK in row_full:
                                    current_node = config.NODE_VEHICLE_CHECK
                                elif config.NODE_BUSINESS_REVIEW in row_full:
                                    current_node = config.NODE_BUSINESS_REVIEW
                                elif config.NODE_VEHICLE_ANNUAL in row_full:
                                    current_node = config.NODE_VEHICLE_ANNUAL
                                elif config.NODE_ARCHIVE in row_full:
                                    current_node = config.NODE_ARCHIVE
                                # 检查是否带挂流程 (车牌含“挂”字也算带挂)
                                if config.FLOW_TRAILER_MARKER in row_full or "(普货)" in row_full or "挂" in t:
                                    flow_name = config.FLOW_TRAILER_MARKER
                                else:
                                    flow_name = config.FLOW_NORMAL_MARKER
                                workbench_plates.append((t, current_node, flow_name))
                                break
                except:
                    pass

                if workbench_plates:
                    # === 第二步: 工作台有任务, 跑完所有任务 ===
                    print(f"\n📋 工作台有 {len(workbench_plates)} 个待办任务:")
                    for p, n, f in workbench_plates:
                        print(f"  · {p} ({f} | {n})")
                    for plate, current_node, flow_name in workbench_plates:
                        if plate in SKIP_PLATES:
                            print(f"  ⏭ 跳过 {plate} (已被拉黑)")
                            continue
                        step_map = {
                            config.NODE_VEHICLE_CHECK: config.STEP_VEHICLE_CHECK,
                            config.NODE_TECH_REVIEW: config.STEP_TECH_REVIEW,
                            config.NODE_BUSINESS_REVIEW: config.STEP_BUSINESS_REVIEW,
                            config.NODE_VEHICLE_ANNUAL: config.STEP_VEHICLE_ANNUAL,
                            config.NODE_ARCHIVE: config.STEP_ARCHIVE,
                        }
                        run_from_step = step_map.get(current_node, "unknown")
                        print(f"\n🔄 续跑: {plate} ({flow_name} | {current_node})")

                        processed += 1
                        print(f"\n{'='*55}")
                        print(f"  第 {processed} 辆: {plate}")
                        try:
                            _t0 = _time.time()
                            dispatch(page, plate, run_from_step, processed)
                            print(f"\n  ✅ {plate} 全部流程完成")
                            RESULTS.append({
                                "plate": plate,
                                "flow_type": "带挂" if "挂" in plate else "不带挂",
                                "start_time": _t0,
                                "end_time": _time.time(),
                                "status": "成功",
                                "error": "",
                            })
                            failures.pop(plate, None)
                        except Exception as e:
                            failures[plate] = failures.get(plate, 0) + 1
                            fc = failures[plate]
                            print(f"\n  ❌ {plate} 处理失败 ({fc}/{config.MAX_FAIL}): {e}")
                            screenshot_on_error(page, plate)
                            _close_all_popups(page)
                            RESULTS.append({
                                "plate": plate,
                                "flow_type": "带挂" if "挂" in plate else "不带挂",
                                "start_time": _t0,
                                "end_time": _time.time(),
                                "status": "失败",
                                "error": str(e),
                            })
                            if fc >= config.MAX_FAIL:
                                SKIP_PLATES.add(plate)
                                print(f"  ⏭ {plate} 已加入黑名单,不再自动重试")
                        pa(1.5)
                    continue  # 跑完所有任务后回到顶部重新检查工作台

                # === 第三步: 工作台空, 去列表取新车 ===
                run_from_step = None
                # 点击货运管理的展开按钮,再点普货审验
                contents = get_contents(page)
                try:
                    # 多重策略点击货运管理展开按钮
                    _click_expand_freight_manage(contents)
                    pa(2)

                    # 点普货审验 (增加文本定位兑底,解决 role 定位超时)
                    _click_normal_review_link(contents)
                    pa(2)

                    # 验证 main_kef 是否真的切到了普货审验列表
                    if not _verify_in_normal_review(page):
                        print(f"  ⚠️ 验证未切到普货审验, 重试点击...")
                        # 再试一次
                        _click_normal_review_link(contents)
                        pa(2)
                except Exception as e:
                    print(f"  普货审验自动导航失败: {e}")
                    safe_input("  >>> 请手动点左侧[普货审验],然后按回车...")
                plate = get_next_plate_from_list(page)
                if not plate:
                    empty_count += 1
                    if empty_count >= 2:
                        print("\n  📋 年审列表和工作台均无待办车辆,流程结束 ✅")
                        break
                    print(f"  ⏳ 未获取到车牌,第{empty_count}次重试...")
                    pa(2)
                    continue
                empty_count = 0

                if plate in SKIP_PLATES:
                    print(f"  ⏭ 跳过 {plate} (已被拉黑)")
                    pa(1)
                    continue

                processed += 1
                print(f"\n{'='*55}")
                print(f"  第 {processed} 辆: {plate}")
                try:
                    _t0 = _time.time()
                    dispatch(page, plate, run_from_step, processed)
                    print(f"\n  ✅ {plate} 全部流程完成")
                    RESULTS.append({
                        "plate": plate,
                        "flow_type": "带挂" if "挂" in plate else "不带挂",
                        "start_time": _t0,
                        "end_time": _time.time(),
                        "status": "成功",
                        "error": "",
                    })
                    failures.pop(plate, None)
                    step("当前车辆完成,准备处理下一辆")
                except Exception as e:
                    failures[plate] = failures.get(plate, 0) + 1
                    fc = failures[plate]
                    print(f"\n  ❌ {plate} 处理失败 ({fc}/{config.MAX_FAIL}): {e}")
                    screenshot_on_error(page, plate)
                    _close_all_popups(page)
                    RESULTS.append({
                        "plate": plate,
                        "flow_type": "带挂" if "挂" in plate else "不带挂",
                        "start_time": _t0,
                        "end_time": _time.time(),
                        "status": "失败",
                        "error": str(e),
                    })
                    if fc >= config.MAX_FAIL:
                        SKIP_PLATES.add(plate)
                        print(f"  ⏭ {plate} 已加入黑名单,不再自动重试")
                pa(1.5)

                if config.SINGLE_RUN:
                    print("\n  🛑 单次模式,已处理 1 辆,停止")
                    break

                if config.MAX_CARS > 0 and processed >= config.MAX_CARS:
                    print(f"\n  🛑 已达上限 {config.MAX_CARS} 辆,停止")
                    break

        except KeyboardInterrupt:
            print("\n  🛑 用户中断")
        except Exception as e:
            print(f"\n  💥 系统异常: {e}")
            screenshot_on_error(page, "fatal")
        finally:
            try:
                for pg in context.pages:
                    try: pg.once("dialog", lambda d: d.accept())
                    except: pass
                context.close()
            except:
                pass
            try:
                browser.close()
            except:
                pass
            # 导出处理结果
            if RESULTS:
                path = export_results_excel()
                if path:
                    print(f"\n  📊 处理结果已导出: {path}")
                    print(f"  📈 总计: {len(RESULTS)} 辆 (成功: {sum(1 for r in RESULTS if r['status']=='成功')}, 失败: {sum(1 for r in RESULTS if r['status']=='失败')})")
            print("  🚪 程序已退出")

if __name__ == "__main__":
    main()
